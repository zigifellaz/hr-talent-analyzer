import streamlit as st
import streamlit.components.v1 as components
import anthropic
import base64
import io
import json
import re
from pathlib import Path
from datetime import datetime

# ─── Page Config ────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Talent Intelligence · M.I.Tech",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── Custom CSS — Editorial Luxury ──────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,600;1,9..40,300&family=Noto+Serif+KR:wght@300;400;600;700&family=Noto+Sans+KR:wght@300;400;500;600&display=swap');

:root {
    --ink:       #1A1714;
    --ink-2:     #3D3830;
    --ink-3:     #7A7268;
    --ink-4:     #B0A898;
    --paper:     #F7F3ED;
    --paper-2:   #EDE8E0;
    --paper-3:   #E2DDD4;
    --gold:      #B8924A;
    --gold-lt:   #D4AF72;
    --slate:     #2B3D5C;
    --slate-lt:  #4A6080;
    --red:       #8B2635;
    --green:     #2D6A4F;
    --rule:      #D4CEC4;
}

/* ── Reset & Base ── */
html, body, .stApp {
    background-color: var(--paper) !important;
    color: var(--ink) !important;
    font-family: 'DM Sans', 'Noto Sans KR', sans-serif !important;
}

/* ── 좌우 여백 ── */
.block-container {
    max-width: 1500px !important;
    padding-left: 3rem !important;
    padding-right: 3rem !important;
    margin-left: auto !important;
    margin-right: auto !important;
}

/* Grain overlay */
.stApp::before {
    content: '';
    position: fixed;
    inset: 0;
    background-image: url("data:image/svg+xml,%3Csvg viewBox='0 0 512 512' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.75' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)' opacity='0.035'/%3E%3C/svg%3E");
    pointer-events: none;
    z-index: 9999;
    opacity: 0.6;
}

/* ── Masthead ── */
.masthead {
    border-bottom: 2px solid var(--ink);
    padding: 1.8rem 0 1.2rem 0;
    margin-bottom: 0;
    display: flex;
    align-items: flex-end;
    justify-content: space-between;
}
.masthead-brand {
    font-family: 'DM Serif Display', 'Noto Serif KR', serif;
    font-size: 2.6rem;
    font-weight: 400;
    color: var(--ink);
    letter-spacing: -1px;
    line-height: 1;
}
.masthead-brand em {
    font-style: italic;
    color: var(--gold);
}
.masthead-meta {
    font-size: 0.7rem;
    color: var(--ink-3);
    text-align: right;
    letter-spacing: 2px;
    text-transform: uppercase;
    line-height: 1.8;
}
.masthead-rule {
    height: 1px;
    background: var(--ink);
    margin: 0.3rem 0 2rem 0;
}
.masthead-thin {
    height: 1px;
    background: var(--rule);
    margin: 0.4rem 0 0 0;
}

/* ── Section Headers ── */
.section-header {
    display: flex;
    align-items: center;
    gap: 1rem;
    margin: 2.2rem 0 1.2rem 0;
}
.section-num {
    font-family: 'DM Serif Display', serif;
    font-size: 0.75rem;
    color: var(--gold);
    letter-spacing: 3px;
    font-style: italic;
    min-width: 24px;
}
.section-title {
    font-size: 0.65rem;
    font-weight: 600;
    letter-spacing: 4px;
    text-transform: uppercase;
    color: var(--ink-2);
    border-bottom: none;
    padding: 0;
    margin: 0;
}
.section-rule {
    flex: 1;
    height: 1px;
    background: var(--rule);
}

/* ── Input Fields ── */
.stTextInput > div > div {
    background: white !important;
    border: 1px solid var(--rule) !important;
    border-radius: 4px !important;
    color: var(--ink) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.9rem !important;
    transition: border-color 0.2s !important;
}
.stTextInput > div > div:focus-within {
    border-color: var(--gold) !important;
    box-shadow: 0 0 0 3px rgba(184,146,74,0.1) !important;
}
.stTextArea > div > div {
    background: white !important;
    border: 1px solid var(--rule) !important;
    border-radius: 4px !important;
    color: var(--ink) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.9rem !important;
}
.stTextArea > div > div:focus-within {
    border-color: var(--gold) !important;
    box-shadow: 0 0 0 3px rgba(184,146,74,0.1) !important;
}
label, .stTextInput label, .stTextArea label {
    font-size: 0.72rem !important;
    font-weight: 600 !important;
    letter-spacing: 1.5px !important;
    text-transform: uppercase !important;
    color: var(--ink-3) !important;
}

/* ── Upload Cards ── */
.upload-item {
    background: white;
    border: 1px solid var(--rule);
    border-radius: 6px;
    padding: 1rem 1.2rem 0.6rem 1.2rem;
    margin-bottom: 0.8rem;
    transition: border-color 0.2s, box-shadow 0.2s;
}
.upload-item:hover {
    border-color: var(--gold);
    box-shadow: 0 2px 12px rgba(184,146,74,0.08);
}
.upload-item-title {
    font-size: 0.78rem;
    font-weight: 600;
    color: var(--ink);
    letter-spacing: 0.5px;
    margin-bottom: 0.15rem;
}
.upload-item-desc {
    font-size: 0.7rem;
    color: var(--ink-4);
    margin-bottom: 0.5rem;
}

/* ── File Uploader ── */
.stFileUploader > div {
    background: var(--paper-2) !important;
    border: 1.5px dashed var(--rule) !important;
    border-radius: 6px !important;
    transition: border-color 0.2s !important;
}
.stFileUploader > div:hover {
    border-color: var(--gold) !important;
}
.stFileUploader label { display: none !important; }

/* ── Analyze Button ── */
.stButton > button {
    background: var(--ink) !important;
    color: var(--paper) !important;
    border: none !important;
    border-radius: 4px !important;
    padding: 0.85rem 2.5rem !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.75rem !important;
    font-weight: 600 !important;
    letter-spacing: 3px !important;
    text-transform: uppercase !important;
    transition: all 0.25s !important;
    width: 100% !important;
    position: relative !important;
    overflow: hidden !important;
}
.stButton > button::before {
    content: '';
    position: absolute;
    inset: 0;
    background: linear-gradient(135deg, var(--gold) 0%, var(--slate) 100%);
    opacity: 0;
    transition: opacity 0.3s;
}
.stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 24px rgba(26,23,20,0.25) !important;
    letter-spacing: 4px !important;
}

/* ── Expander ── */
div[data-testid="stExpander"] {
    background: white !important;
    border: 1px solid var(--rule) !important;
    border-radius: 6px !important;
}
.streamlit-expanderHeader {
    font-size: 0.75rem !important;
    font-weight: 600 !important;
    letter-spacing: 1px !important;
    color: var(--ink-2) !important;
    text-transform: uppercase !important;
}

/* ── Result: Report Cover ── */
.report-cover {
    background: var(--ink);
    border-radius: 8px;
    padding: 3rem 3.5rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
}
.report-cover::after {
    content: '◈';
    position: absolute;
    right: 3rem;
    top: 50%;
    transform: translateY(-50%);
    font-size: 8rem;
    color: rgba(255,255,255,0.04);
    line-height: 1;
    pointer-events: none;
}
.report-cover-label {
    font-size: 0.62rem;
    font-weight: 600;
    letter-spacing: 4px;
    text-transform: uppercase;
    color: var(--gold);
    margin-bottom: 1rem;
}
.report-cover-name {
    font-family: 'DM Serif Display', 'Noto Serif KR', serif;
    font-size: 3rem;
    font-weight: 400;
    color: var(--paper);
    line-height: 1.1;
    margin-bottom: 0.8rem;
    letter-spacing: -1px;
}
.report-cover-summary {
    font-size: 0.95rem;
    color: rgba(247,243,237,0.65);
    line-height: 1.7;
    font-weight: 300;
    max-width: 580px;
    margin-bottom: 1.5rem;
}
.tag-chip {
    display: inline-block;
    border: 1px solid rgba(184,146,74,0.5);
    color: var(--gold-lt);
    border-radius: 3px;
    padding: 0.25rem 0.7rem;
    font-size: 0.68rem;
    font-weight: 500;
    letter-spacing: 1px;
    margin: 0.2rem 0.2rem 0 0;
    text-transform: uppercase;
}

/* ── Dimension Cards ── */
.dim-card {
    background: white;
    border: 1px solid var(--rule);
    border-radius: 6px;
    padding: 1.6rem 1.8rem;
    margin-bottom: 1rem;
    position: relative;
    overflow: hidden;
    transition: box-shadow 0.2s;
}
.dim-card:hover {
    box-shadow: 0 4px 20px rgba(26,23,20,0.08);
}
.dim-card::before {
    content: '';
    position: absolute;
    left: 0; top: 0; bottom: 0;
    width: 3px;
    background: var(--gold);
}
.dim-header {
    display: flex;
    align-items: flex-start;
    justify-content: space-between;
    margin-bottom: 1rem;
}
.dim-icon-title {
    display: flex;
    align-items: center;
    gap: 0.6rem;
}
.dim-icon {
    width: 32px;
    height: 32px;
    background: var(--paper-2);
    border-radius: 6px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 1rem;
}
.dim-name {
    font-size: 0.8rem;
    font-weight: 600;
    color: var(--ink);
    letter-spacing: 0.5px;
}
.dim-sub {
    font-size: 0.65rem;
    color: var(--ink-4);
    letter-spacing: 1px;
    text-transform: uppercase;
}
.dim-score-block {
    text-align: right;
}
.dim-score {
    font-family: 'DM Serif Display', serif;
    font-size: 2.4rem;
    color: var(--ink);
    line-height: 1;
    font-style: italic;
}
.dim-score span {
    font-size: 0.9rem;
    color: var(--ink-4);
    font-style: normal;
    font-family: 'DM Sans', sans-serif;
}
.dim-grade {
    font-size: 0.65rem;
    font-weight: 700;
    letter-spacing: 2px;
    margin-top: 0.2rem;
    text-transform: uppercase;
}
.progress-track {
    height: 3px;
    background: var(--paper-3);
    border-radius: 999px;
    margin: 0.8rem 0 1rem 0;
    overflow: hidden;
}
.progress-fill {
    height: 100%;
    border-radius: 999px;
    background: linear-gradient(90deg, var(--gold), var(--slate));
}
.dim-summary {
    font-size: 0.83rem;
    color: var(--ink-2);
    line-height: 1.75;
    margin-bottom: 0.8rem;
}
.evidence-list {
    display: flex;
    flex-direction: column;
    gap: 0.3rem;
    padding-top: 0.8rem;
    border-top: 1px solid var(--paper-3);
}
.evidence-item {
    display: flex;
    align-items: flex-start;
    gap: 0.5rem;
    font-size: 0.75rem;
    color: var(--ink-3);
    line-height: 1.5;
}
.evidence-dot {
    width: 4px;
    height: 4px;
    background: var(--gold);
    border-radius: 50%;
    margin-top: 0.45rem;
    flex-shrink: 0;
}

/* ── Keyword Cards ── */
.kw-card {
    background: white;
    border: 1px solid var(--rule);
    border-radius: 6px;
    padding: 1.8rem 2rem;
    margin-bottom: 1rem;
    display: grid;
    grid-template-columns: 56px 1fr;
    gap: 1.5rem;
    align-items: start;
    transition: box-shadow 0.2s;
}
.kw-card:hover {
    box-shadow: 0 4px 20px rgba(26,23,20,0.08);
}
.kw-rank-col {
    text-align: center;
    padding-top: 0.2rem;
}
.kw-rank-num {
    font-family: 'DM Serif Display', serif;
    font-size: 2.8rem;
    font-style: italic;
    line-height: 1;
}
.rank-gold { color: #B8924A; }
.rank-silver { color: #8A9BA8; }
.rank-bronze { color: #9B7B5A; }
.kw-rank-label {
    font-size: 0.6rem;
    letter-spacing: 2px;
    text-transform: uppercase;
    color: var(--ink-4);
    display: block;
    margin-top: 0.2rem;
}
.kw-title {
    font-family: 'DM Serif Display', 'Noto Serif KR', serif;
    font-size: 1.35rem;
    color: var(--ink);
    margin-bottom: 0.5rem;
    line-height: 1.2;
}
.kw-why {
    font-size: 0.82rem;
    color: var(--ink-2);
    line-height: 1.75;
    margin-bottom: 0.9rem;
}
.kw-how {
    background: var(--paper-2);
    border-radius: 4px;
    padding: 0.8rem 1rem;
    border-left: 2px solid var(--gold);
}
.kw-how-label {
    font-size: 0.6rem;
    font-weight: 700;
    letter-spacing: 3px;
    text-transform: uppercase;
    color: var(--gold);
    margin-bottom: 0.4rem;
}
.kw-how-text {
    font-size: 0.8rem;
    color: var(--ink-2);
    line-height: 1.65;
}

/* ── Insight Box ── */
.insight-box {
    background: var(--slate);
    border-radius: 8px;
    padding: 2.5rem 3rem;
    position: relative;
    overflow: hidden;
}
.insight-box::before {
    content: '"';
    position: absolute;
    left: 2rem;
    top: -1rem;
    font-family: 'DM Serif Display', serif;
    font-size: 10rem;
    color: rgba(255,255,255,0.06);
    line-height: 1;
    pointer-events: none;
}
.insight-label {
    font-size: 0.6rem;
    font-weight: 700;
    letter-spacing: 4px;
    text-transform: uppercase;
    color: rgba(212,175,114,0.8);
    margin-bottom: 1rem;
}
.insight-text {
    font-family: 'DM Sans', 'Noto Sans KR', sans-serif;
    font-size: 0.95rem;
    color: rgba(247,243,237,0.85);
    line-height: 2;
    font-weight: 300;
}

/* ── Horizontal Rule ── */
.hr {
    height: 1px;
    background: var(--rule);
    border: none;
    margin: 2.5rem 0;
}

/* ── Success / Error ── */
.stSuccess { background: #EAF4EE !important; border-color: var(--green) !important; }
.stError { background: #FAEAEC !important; border-color: var(--red) !important; }

/* ── 새 분석 시작 버튼 ── */
.new-analysis-btn > div > button {
    background: white !important;
    color: var(--ink) !important;
    border: 1.5px solid var(--gold) !important;
    color: var(--gold) !important;
}
</style>
""", unsafe_allow_html=True)


# ─── Helpers ────────────────────────────────────────────────────────────────
def read_file_content(uploaded_file) -> str:
    name = uploaded_file.name.lower()
    content = uploaded_file.read()
    if name.endswith(".pdf"):
        try:
            import pdfplumber
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                return "\n".join(p.extract_text() or "" for p in pdf.pages)
        except Exception:
            pass
        try:
            import PyPDF2
            reader = PyPDF2.PdfReader(io.BytesIO(content))
            return "\n".join(p.extract_text() or "" for p in reader.pages)
        except Exception:
            return "[PDF — 텍스트 추출 실패]"
    if name.endswith(".docx"):
        try:
            import docx
            doc = docx.Document(io.BytesIO(content))
            return "\n".join(p.text for p in doc.paragraphs)
        except Exception:
            return "[DOCX — 텍스트 추출 실패]"
    if any(name.endswith(e) for e in [".jpg", ".jpeg", ".png", ".webp", ".gif"]):
        ext_mime = {
            ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
            ".png": "image/png", ".webp": "image/webp",
            ".gif": "image/gif"
        }
        mime = next((v for k, v in ext_mime.items() if name.endswith(k)), "image/jpeg")

        # 이미지 자동 압축 (Claude API 권장: 1568px 이하)
        try:
            from PIL import Image as PILImage
            img = PILImage.open(io.BytesIO(content))
            # RGBA/P 모드 → RGB 변환 (JPEG 저장 필요)
            if img.mode in ("RGBA", "P", "LA", "CMYK"):
                img = img.convert("RGB")
                mime = "image/jpeg"
            # 최대 1568px 리사이즈
            max_px = 1568
            w, h = img.size
            if w > max_px or h > max_px:
                ratio = min(max_px / w, max_px / h)
                img = img.resize((int(w * ratio), int(h * ratio)), PILImage.LANCZOS)
            # 압축 저장
            buf = io.BytesIO()
            if mime == "image/jpeg":
                img.save(buf, format="JPEG", quality=82, optimize=True)
            else:
                img = img.convert("RGB")
                img.save(buf, format="JPEG", quality=82, optimize=True)
                mime = "image/jpeg"
            content = buf.getvalue()
        except Exception:
            pass  # 압축 실패 시 원본 그대로 사용

        b64 = base64.standard_b64encode(content).decode()
        return "__IMGOBJ__" + json.dumps({"mime": mime, "b64": b64})
    try:
        return content.decode("utf-8")
    except Exception:
        try:
            return content.decode("cp949")
        except Exception:
            return "[파일 읽기 실패]"


def build_user_content(file_data, candidate_name, company_standard):
    user_content = []
    text_parts = []
    if candidate_name:
        text_parts.append(f"[대상자 이름] {candidate_name}")
    if company_standard:
        text_parts.append(f"[회사 인재상]\n{company_standard}")
    for label, content in file_data.items():
        if content and not content.startswith("__IMGOBJ__"):
            text_parts.append(f"[{label}]\n{content}")
    if text_parts:
        user_content.append({"type": "text", "text": "\n\n".join(text_parts)})
    for label, content in file_data.items():
        if content and content.startswith("__IMGOBJ__"):
            try:
                img = json.loads(content[len("__IMGOBJ__"):])
                mime_type = img.get("mime", "image/jpeg")
                b64_data  = img.get("b64", "")
                valid_mimes = ["image/jpeg", "image/png", "image/webp", "image/gif"]
                if mime_type not in valid_mimes:
                    mime_type = "image/jpeg"
                if not b64_data:
                    continue
                user_content.append({
                    "type": "image",
                    "source": {"type": "base64", "media_type": mime_type, "data": b64_data}
                })
                user_content.append({"type": "text", "text": f"위 이미지는 [{label}] 자료입니다."})
            except Exception:
                pass
    return user_content


def analyze_candidate(api_key, file_data, candidate_name, company_standard):
    client = anthropic.Anthropic(api_key=api_key)
    system_prompt = """당신은 글로벌 탑티어 HR 컨설팅 펌(McKinsey People & Organization, Korn Ferry, Spencer Stuart 수준)의 수석 어세스먼트 컨설턴트입니다. 조직심리학 박사 학위와 15년 이상의 임원 평가 및 인재 어세스먼트 경험을 보유하고 있습니다.

【핵심 분석 목적 — 조직 리밸런싱 (Rebalancing)】
본 분석의 궁극적 목적은 조직 재구성을 위한 의사결정 지원입니다. 구체적으로:
1. 누가 조직에 핵심적으로 남아야 하는가(Keep)
2. 누가 조직 방향성과 적합하지 않은가(Misfit)
3. 누구에게 리더 역할을 부여할 수 있는가(Leadership Readiness)
4. 각 인원의 향후 커리어 트랙(Career Track) 방향
이 4가지 판단에 직결되는 실용적 인사이트를 제공해야 합니다.

당신의 분석은 다음 프레임워크를 통합적으로 적용합니다:
- Korn Ferry의 Leadership Architect (역량 모델 67개 팩터)
- SHL의 OPQ32 (성격 및 행동 선호도 측정)
- Hogan Assessment의 HPI/HDS/MVPI (명시적 성격 / 암묵적 위험 요소 / 동기 가치 체계)
- DDI의 Targeted Selection (행동사건 면접법 기반 역량 평가)
- MBTI 및 Big Five(OCEAN) 모델과의 교차 검증

━━━━━━━━━━━━━━━━━━━━━━━━
분석 원칙
━━━━━━━━━━━━━━━━━━━━━━━━
1. 근거 기반 추론: 모든 평가는 제공된 자료에서 직접 인용 가능한 구체적 근거를 바탕으로 합니다. 추측성 표현("~할 것 같다") 대신 행동 증거 기반 표현("~한 이력이 확인된다")을 사용합니다.
2. 다층적 교차 검증: 단일 자료가 아닌 복수 자료 간 일관성·불일치를 분석하여 표면 행동과 내재 동기를 구분합니다.
3. 조직 적합도 연계: 개인 역량 분석을 회사 인재상 및 직무 요구사항과 명시적으로 연결합니다.
4. 위험 요인 식별: 강점 이면의 잠재적 취약점(Derailer)을 전문가 시각으로 도출합니다.
5. 리밸런싱 판단 직결: 분석 결과를 Keep/Misfit/리더 적합성/커리어 트랙 판단으로 명확히 전환합니다.
6. SNS 자료 활용: SNS 자료가 제공된 경우, 개인 성향 및 내부 조직문화 적합성 파악 목적으로만 분석하며 대외비로 취급합니다.

━━━━━━━━━━━━━━━━━━━━━━━━
추가 정량 평가 항목 (조직 적합도 & 리더십 준비도)
━━━━━━━━━━━━━━━━━━━━━━━━
[조직 적합도 - Organizational Fit] (100점 만점)
- 인재상 부합도: 회사 인재상(성장지향/상호존중/혁신과 도전)과의 일치도
- 문화 적합성: 5대 핵심문화 축(개방적 소통/몰입 실행/성과 인정/협업 시너지/혁신 리더십)과의 정합성
- 방향성 일치: 엠아이텍의 향후 전략 방향과 개인 역량·성향의 정렬도
- 점수가 낮을수록(50점 미만) 조직 부적합(Misfit) 신호로 해석

[리더십 준비도 - Leadership Readiness] (100점 만점)
- 현재 리더 역할 수행 가능 수준을 정량 평가
- 80점 이상: 즉시 리더 역할 부여 가능 / 60-79: 육성 후 부여 / 60점 미만: 리더 역할 부적합
- 리더십 잠재력과 현재 발현 수준을 구분하여 평가

[학력-성과 정합성 - Credential-Performance Alignment]
- 학력 수준(학위/대학) 대비 실제 업무 성과의 일치 여부를 분석
- 고학력이 실제 성과로 이어지는지, 또는 학력 대비 성과가 미흡한지 판별

━━━━━━━━━━━━━━━━━━━━━━━━
각 차원별 평가 기준
━━━━━━━━━━━━━━━━━━━━━━━━
[인지 능력 - Cognitive Ability]
- 개념적 사고력: 복잡한 정보를 구조화하고 패턴을 도출하는 능력
- 분석적 추론: 데이터/상황에서 핵심 변수를 식별하고 인과관계를 파악하는 능력
- 학습 민첩성 (Learning Agility): 새로운 환경과 정보에 빠르게 적응하는 능력
- 의사결정 질: 불확실한 상황에서 논리적이고 신속한 판단을 내리는 능력

[잡 전문성 - Job Expertise]
- 직무 지식 깊이: 해당 산업/직무의 핵심 지식 및 기술 수준
- 실행 역량: 지식을 실제 성과로 전환하는 능력 (KPI 달성 이력 포함)
- 도메인 네트워크: 업계 내 관계망과 시장 이해도
- 글로벌 역량: 크로스컬처 환경에서의 협업·소통 능력

[적극성 - Proactiveness]
- 주도성 (Initiative): 지시 없이 과제를 발굴하고 선제적으로 행동하는 성향
- 결과 지향성 (Achievement Drive): 목표 달성에 대한 내적 동기 강도
- 변화 주도: 현상 유지보다 개선과 혁신을 추구하는 성향
- 역경 극복 (Resilience): 실패와 장애 상황에서의 회복탄력성

[리더십 - Leadership]
- 영향력 행사: 공식 권한 없이도 타인을 설득하고 이끄는 능력
- 팀 개발: 구성원의 성장을 지원하고 동기를 부여하는 능력
- 전략적 방향 설정: 조직의 장기 비전을 수립하고 전달하는 능력
- 이해관계자 관리: 내외부 이해관계자와의 관계를 전략적으로 구축하는 능력

━━━━━━━━━━━━━━━━━━━━━━━━
점수 산출 기준 (100점 만점)
━━━━━━━━━━━━━━━━━━━━━━━━
90-100 (S): 동종업계 상위 5% 수준. 해당 역량의 롤모델.
80-89 (A): 상위 15% 수준. 명확한 강점으로 조직에 즉각적 기여 가능.
70-79 (B+): 상위 30% 수준. 강점이 있으나 일부 개발 영역 존재.
60-69 (B): 평균 수준. 기본 역량은 갖추었으나 차별화 요소 미흡.
50-59 (B-): 평균 이하. 해당 역량에서 주의 깊은 관찰과 개발 지원 필요.
49 이하 (C): 유의미한 약점. 해당 역량이 직무 핵심 요건이라면 채용 리스크.

━━━━━━━━━━━━━━━━━━━━━━━━
번아웃 위험도 평가 기준 (Maslach Burnout Inventory 모델 기반)
━━━━━━━━━━━━━━━━━━━━━━━━
번아웃 위험도는 아래 3개 축을 종합하여 LOW / MEDIUM / HIGH / CRITICAL 4단계로 평가합니다:
- 정서적 고갈 (Emotional Exhaustion): 에너지 소진, 감정적 탈진 신호
- 비인격화 (Depersonalization): 냉소적 태도, 직무 의미 상실 신호
- 개인 성취감 저하 (Reduced Personal Accomplishment): 무력감, 자기 효능감 하락 신호
근거 자료: SNS 어조, 다면평가 결과, 기안서 문체, MBTI 스트레스 반응 패턴 등을 교차 분석

━━━━━━━━━━━━━━━━━━━━━━━━
이직 가능성 평가 기준 (Push-Pull 모델 기반)
━━━━━━━━━━━━━━━━━━━━━━━━
이직 가능성은 아래 Push(현조직 이탈 요인)와 Pull(외부 유인 요인)을 종합하여 LOW / MEDIUM / HIGH / CRITICAL 4단계로 평가합니다:
- Push 요인: 성장 정체감, 보상 불만족 신호, 관계 갈등, 번아웃 수준
- Pull 요인: 외부 네트워크 활동성, 스킬 시장가치, 업계 이동성
- 재직 의향 신호: 장기 프로젝트 참여도, 조직 애착 언어, 커리어 방향성
근거 자료: 이력서 재직 기간 패턴, SNS 활동, 다면평가 몰입도, 기안서 미래 지향성 등 교차 분석

━━━━━━━━━━━━━━━━━━━━━━━━
추가 분석 항목 (신규)
━━━━━━━━━━━━━━━━━━━━━━━━
[프로필 구조화 추출 - Profile]
- 제공된 자료에서 전공(major), 출신 대학(university), 최종 학력 수준(education_level), 출신 지역(region)을 추출합니다.
- education_level은 '박사 / 석사 / 학사 / 전문학사 / 고졸 / 자료 미제공' 중 하나로 정규화합니다.
- 확인되지 않는 항목은 '자료 미제공'으로 표기합니다. 추측하지 마십시오.

[역량 세부 점수 - Sub-scores]
- 4대 역량 각각에 대해, 위에 정의된 4개 세부 기준을 0~100으로 각각 점수화합니다.
  · 인지 능력: 개념적사고 / 분석적추론 / 학습민첩성 / 의사결정질
  · 잡 전문성: 직무지식깊이 / 실행역량 / 도메인네트워크 / 글로벌역량
  · 적극성: 주도성 / 결과지향성 / 변화주도 / 역경극복
  · 리더십: 영향력행사 / 팀개발 / 전략적방향설정 / 이해관계자관리

[회사 향후 방향성 적합도 - Direction Fit] (100점 만점)
- '[회사 향후 방향성]' 자료가 제공된 경우, 그 전략 방향과 개인 역량·성향의 정렬도를 별도 점수로 평가합니다.
- 해당 자료가 없으면 인재상·핵심문화를 근거로 보수적으로 추정하고 그 사실을 summary에 명시합니다.

[SNS 전용 분석 - SNS Analysis] (대외비)
- SNS 자료가 제공된 경우에만 available=true로 설정하고, 개인 성향(personality)과 내부 조직문화 적합성(culture_fit)을 분석하며 0~100 점수를 부여합니다.
- SNS 자료가 없으면 available=false로 설정하고 나머지는 빈 문자열로 둡니다. 이 항목은 항상 대외비(confidential=true)입니다.

[분석 자료 커버리지 - Data Coverage]
- 분석에 실제로 활용된 자료 유형(이력서/기안서/MBTI/인적성/SNS/다면평가/기타)을 나열하고, 근거의 충분성을 HIGH/MEDIUM/LOW로 평가합니다.

반드시 아래 JSON 형식으로만 응답하세요. JSON 외 어떤 텍스트도 출력하지 마세요:
{
  "candidate_summary": "대상자 핵심 특성 한줄 평가 (50자 이내)",
  "personality_tags": ["태그1","태그2","태그3","태그4","태그5"],
  "profile": {
    "major": "전공 (없으면 '자료 미제공')",
    "university": "출신 대학 (없으면 '자료 미제공')",
    "education_level": "박사 / 석사 / 학사 / 전문학사 / 고졸 / 자료 미제공 중 택1",
    "region": "출신 지역 (없으면 '자료 미제공')"
  },
  "data_coverage": {
    "materials": ["활용된 자료 유형 나열"],
    "confidence": "HIGH / MEDIUM / LOW 중 택1",
    "note": "1문장. 분석 근거 충분성 평가"
  },
  "dimensions": {
    "cognitive_ability": {
      "score": 75,
      "grade": "B+",
      "sub_scores": {"개념적사고": 78, "분석적추론": 72, "학습민첩성": 80, "의사결정질": 70},
      "summary": "3문장 분석. 강점 발현 방식·조직 활용 가능성·잠재 한계 포함",
      "evidence": ["근거1","근거2"]
    },
    "job_expertise": {
      "score": 80,
      "grade": "A",
      "sub_scores": {"직무지식깊이": 82, "실행역량": 80, "도메인네트워크": 76, "글로벌역량": 78},
      "summary": "3문장 분석",
      "evidence": ["근거1","근거2"]
    },
    "proactiveness": {
      "score": 70,
      "grade": "B",
      "sub_scores": {"주도성": 72, "결과지향성": 70, "변화주도": 68, "역경극복": 70},
      "summary": "3문장 분석",
      "evidence": ["근거1","근거2"]
    },
    "leadership": {
      "score": 65,
      "grade": "B-",
      "sub_scores": {"영향력행사": 66, "팀개발": 64, "전략적방향설정": 65, "이해관계자관리": 66},
      "summary": "3문장 분석",
      "evidence": ["근거1","근거2"]
    }
  },
  "burnout_risk": {
    "level": "MEDIUM",
    "score": 45,
    "emotional_exhaustion": "2문장",
    "depersonalization": "2문장",
    "personal_accomplishment": "2문장",
    "summary": "2문장 총평 및 관리 권고",
    "evidence": ["근거1","근거2"]
  },
  "turnover_risk": {
    "level": "LOW",
    "score": 25,
    "push_factors": "2문장",
    "pull_factors": "2문장",
    "retention_signals": "2문장",
    "summary": "2문장 총평 및 리텐션 전략",
    "evidence": ["근거1","근거2"]
  },
  "sns_analysis": {
    "available": false,
    "score": 0,
    "personality": "SNS 자료가 있으면 2문장, 없으면 빈 문자열",
    "culture_fit": "SNS 자료가 있으면 2문장, 없으면 빈 문자열",
    "confidential": true
  },
  "hiring_keywords": [
    {
      "rank": 1,
      "keyword": "10자 이내 키워드",
      "why": "2문장 선정 이유 (인재상 연결)",
      "how_to_check": "STAR 기반 질문 1개 + 평가 포인트"
    },
    {
      "rank": 2,
      "keyword": "10자 이내 키워드",
      "why": "2문장",
      "how_to_check": "STAR 기반 질문 1개 + 평가 포인트"
    },
    {
      "rank": 3,
      "keyword": "10자 이내 키워드",
      "why": "2문장",
      "how_to_check": "STAR 기반 질문 1개 + 평가 포인트"
    }
  ],
  "org_fit": {
    "score": 75,
    "grade": "B+",
    "talent_match": "인재상 부합도 1-2문장 평가",
    "culture_fit": "5대 핵심문화 축 정합성 1-2문장 평가",
    "direction_alignment": "회사 방향성 정렬도 1-2문장 평가",
    "summary": "2문장. 조직 적합도 종합 및 Misfit 여부 판단"
  },
  "direction_fit": {
    "score": 70,
    "summary": "2문장. 회사 향후 방향성과 개인의 정렬도. 방향성 자료가 없으면 추정임을 명시"
  },
  "leadership_readiness": {
    "score": 65,
    "level": "육성 후 부여",
    "rationale": "2문장. 리더십 준비도 판단 근거",
    "recommendation": "리더 역할 부여 가능 여부 명확한 결론 (즉시 가능 / 육성 후 가능 / 부적합 중 택1 + 1문장 설명)"
  },
  "credential_performance": {
    "alignment": "일치 / 학력우위 / 성과우위 중 택1",
    "education_level": "확인된 학력 수준 (자료 없으면 '자료 미제공')",
    "summary": "2문장. 학력 대비 실제 성과 정합성 분석"
  },
  "career_track": {
    "current_position": "현재 포지션 추정",
    "recommended_track": "추천 커리어 트랙 (전문가형/관리자형/전환필요 등)",
    "summary": "2-3문장. 향후 커리어 방향 제언"
  },
  "rebalancing_verdict": {
    "decision": "KEEP / DEVELOP / WATCH / MISFIT 중 택1",
    "confidence": "HIGH / MEDIUM / LOW 중 택1",
    "rationale": "2-3문장. 리밸런싱 관점 핵심 판단 근거. 조직에 남아야 하는지/방향성과 맞는지 명확히"
  },
  "derailer": "2문장. 스트레스·장기 재직 시 부정적 행동 패턴",
  "development_suggestion": "2문장. 최고 성과를 위한 환경·관리 방식",
  "overall_insight": "4문장. 인재 유형 명명·최적 포지셔닝·리밸런싱 최종 권고"
}"""

    user_content = build_user_content(file_data, candidate_name, company_standard)
    if not user_content:
        raise ValueError("분석할 자료가 없습니다.")
    user_content.append({"type": "text", "text": "위 자료를 바탕으로 조직 리밸런싱(Rebalancing) 관점의 전문 인재 분석을 JSON 형식으로 수행해주세요. 모든 평가는 제공된 자료의 구체적 근거에 기반해야 하며, 특히 '이 인원이 조직에 남아야 하는지', '회사 방향성과 맞는지', '리더 역할 부여가 가능한지', '향후 커리어 트랙은 무엇인지'에 대한 명확한 판단을 제공해주세요."})

    response = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=16000,
        system=system_prompt,
        messages=[{"role": "user", "content": user_content}]
    )
    raw = response.content[0].text.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    # JSON 블록만 추출
    match = re.search(r'\{[\s\S]+\}', raw)
    if match:
        raw = match.group(0)
    # 잘린 JSON 자동 복구: 마지막 완성된 키까지만 살려서 닫기
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        # 마지막 완전한 key-value 쌍 이후를 제거하고 닫기 시도
        repaired = raw
        for closing in ['"}', '"]', '}']:
            last = repaired.rfind(closing)
            if last != -1:
                candidate = repaired[:last + len(closing)]
                # 열린 괄호 수에 맞춰 닫기
                opens  = candidate.count('{') - candidate.count('}')
                opens2 = candidate.count('[') - candidate.count(']')
                candidate += ']' * max(0, opens2) + '}' * max(0, opens)
                try:
                    return json.loads(candidate)
                except Exception:
                    continue
        raise


def grade_color(grade):
    return {"A+":"#2D6A4F","A":"#2D6A4F","B+":"#2B3D5C","B":"#2B3D5C",
            "B-":"#8B6914","C+":"#8B2635","C":"#8B2635"}.get(grade, "#7A7268")

def dim_label(k):
    return {"cognitive_ability":"인지 능력","job_expertise":"잡 전문성",
            "proactiveness":"적극성","leadership":"리더십"}.get(k, k)

def dim_sublabel(k):
    return {"cognitive_ability":"Cognitive Ability","job_expertise":"Job Expertise",
            "proactiveness":"Proactiveness","leadership":"Leadership"}.get(k, k)

def dim_icon(k):
    return {"cognitive_ability":"🧠","job_expertise":"⚙️",
            "proactiveness":"🔥","leadership":"👑"}.get(k,"◈")


# ─── 종합 점수(정량화) ──────────────────────────────────────────
# 종합 점수 가중치 (합계 1.0). 비중을 바꾸고 싶으면 이 숫자만 수정하면 됩니다.
#   talent              : 4대 역량 평균 (인지·전문성·적극성·리더십)
#   org_fit             : 조직 적합도
#   leadership_readiness: 리더십 준비도
#   low_risk            : (100 - 번아웃·이직 평균 위험) — 위험이 낮을수록 점수 상승
OVERALL_WEIGHTS = {
    "talent":               0.35,
    "org_fit":              0.30,
    "leadership_readiness": 0.15,
    "low_risk":             0.20,
}

def _safe_num(v):
    return v if isinstance(v, (int, float)) else 0

def overall_grade(score):
    if score is None: return "-"
    if score >= 90: return "S"
    if score >= 80: return "A"
    if score >= 70: return "B+"
    if score >= 60: return "B"
    if score >= 50: return "B-"
    return "C"

def overall_color(score):
    if score is None: return "#7A7268"
    if score >= 80: return "#2D6A4F"
    if score >= 60: return "#2B3D5C"
    if score >= 50: return "#8B6914"
    return "#8B2635"

def compute_overall_score(R: dict):
    """이미 계산된 세부 점수들을 가중 합산해 0~100 종합 점수를 산출한다.
    저장된 기존 분석에도 그대로 적용되므로 재분석이 필요 없다.
    반환: (점수:int 또는 None, 등급:str)"""
    if not R:
        return None, "-"
    dims = R.get("dimensions", {})
    dvals = [_safe_num(dims.get(k, {}).get("score")) for k in
             ("cognitive_ability", "job_expertise", "proactiveness", "leadership")]
    talent = sum(dvals) / len(dvals) if dvals else 0
    org_fit = _safe_num(R.get("org_fit", {}).get("score"))
    lr      = _safe_num(R.get("leadership_readiness", {}).get("score"))
    burn    = _safe_num(R.get("burnout_risk", {}).get("score"))
    turn    = _safe_num(R.get("turnover_risk", {}).get("score"))
    # 세부 점수가 하나도 없으면(분석 실패 등) 종합 점수도 없음
    if talent == 0 and org_fit == 0 and lr == 0:
        return None, "-"
    risk_avg = (burn + turn) / 2
    w = OVERALL_WEIGHTS
    score = (w["talent"] * talent
             + w["org_fit"] * org_fit
             + w["leadership_readiness"] * lr
             + w["low_risk"] * (100 - risk_avg))
    score = int(max(0, min(100, round(score))))
    return score, overall_grade(score)


# ─── 아카이브 함수 (Supabase 영구 저장 + 로컬 폴백) ─────────────────────────
ARCHIVE_FILE = "archive.json"
SUPABASE_TABLE = "talent_archive"

def _get_supabase():
    """Supabase 연결 정보 반환. 설정 안 됐으면 None."""
    try:
        url = st.secrets.get("SUPABASE_URL", "")
        key = st.secrets.get("SUPABASE_KEY", "")
        if url and key:
            return url.rstrip("/"), key
    except Exception:
        pass
    return None, None

def _sb_headers(key):
    return {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": "return=representation"
    }

def load_archive() -> list:
    """Supabase → 로컬 파일 순으로 시도"""
    import requests as req
    url, key = _get_supabase()
    if url and key:
        try:
            r = req.get(
                f"{url}/rest/v1/{SUPABASE_TABLE}?order=created_at.desc&limit=200",
                headers=_sb_headers(key), timeout=5
            )
            if r.status_code == 200:
                rows = r.json()
                return [
                    {
                        "id":             row.get("id"),
                        "saved_at":       row.get("saved_at", ""),
                        "candidate_name": row.get("candidate_name", ""),
                        "dept":           row.get("dept", ""),
                        "result":         json.loads(row.get("result_json", "{}"))
                    }
                    for row in rows
                ]
        except Exception:
            pass

    # 로컬 폴백
    try:
        if Path(ARCHIVE_FILE).exists():
            with open(ARCHIVE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return []

def save_to_archive(record: dict):
    """Supabase → 로컬 파일 순으로 저장"""
    import requests as req
    url, key = _get_supabase()
    if url and key:
        try:
            payload = {
                "saved_at":       record.get("saved_at", ""),
                "candidate_name": record.get("candidate_name", ""),
                "dept":           record.get("dept", ""),
                "result_json":    json.dumps(record.get("result", {}), ensure_ascii=False)
            }
            r = req.post(
                f"{url}/rest/v1/{SUPABASE_TABLE}",
                headers=_sb_headers(key),
                json=payload, timeout=5
            )
            if r.status_code in (200, 201):
                return  # Supabase 저장 성공
        except Exception:
            pass

    # 로컬 폴백
    archive = load_archive()
    archive.insert(0, record)
    with open(ARCHIVE_FILE, "w", encoding="utf-8") as f:
        json.dump(archive, f, ensure_ascii=False, indent=2)

def delete_from_archive(record_id, idx: int):
    """Supabase row id로 삭제 → 없으면 로컬 인덱스 삭제"""
    import requests as req
    url, key = _get_supabase()
    if url and key and record_id:
        try:
            req.delete(
                f"{url}/rest/v1/{SUPABASE_TABLE}?id=eq.{record_id}",
                headers=_sb_headers(key), timeout=5
            )
            return
        except Exception:
            pass

    # 로컬 폴백
    archive = load_archive()
    if 0 <= idx < len(archive):
        archive.pop(idx)
    with open(ARCHIVE_FILE, "w", encoding="utf-8") as f:
        json.dump(archive, f, ensure_ascii=False, indent=2)


# ─── 조직도 데이터 & 상태 판정 ────────────────────────────────────────────────
def load_org_data() -> dict:
    """org_data.json 로드"""
    try:
        if Path("org_data.json").exists():
            with open("org_data.json", "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def get_person_status(name: str, archive_by_name: dict) -> tuple:
    """
    분석 결과 유무 + 위험도에 따라 상태등 색상 반환
    returns: (status, color, result_or_none)
    status: 'none'(분석없음) / 'green'(정상) / 'yellow'(주의) / 'red'(긴급)
    """
    rec = archive_by_name.get(name)
    if not rec:
        return ("none", "#C0BCB4", None)

    R = rec.get("result", {})
    b_lvl = R.get("burnout_risk", {}).get("level", "LOW")
    t_lvl = R.get("turnover_risk", {}).get("level", "LOW")
    verdict = R.get("rebalancing_verdict", {}).get("decision", "")

    # 긴급(빨강): CRITICAL 위험 또는 MISFIT 판정
    if b_lvl == "CRITICAL" or t_lvl == "CRITICAL" or verdict == "MISFIT":
        return ("red", "#C0392B", R)
    # 주의(노랑): HIGH/MEDIUM 위험 또는 WATCH 판정
    if b_lvl in ("HIGH", "MEDIUM") or t_lvl in ("HIGH", "MEDIUM") or verdict == "WATCH":
        return ("yellow", "#E0A800", R)
    # 정상(초록)
    return ("green", "#2D6A4F", R)


# ─── 결과 렌더링 함수 (신규 분석 & 아카이브 조회 공용) ──────────────────────
def render_result(R: dict, candidate_name: str):
    name_d = candidate_name or "대상자"
    ov_score, ov_grade = compute_overall_score(R)
    tags_html = "".join(
        f'<span class="tag-chip">{t}</span>'
        for t in R.get("personality_tags", [])
    )
    st.markdown(f"""
    <div class="report-cover">
        <div class="report-cover-label">◈ Talent Analysis Report · Rebalancing</div>
        <div class="report-cover-name">{name_d}</div>
        <div class="report-cover-summary">{R.get('candidate_summary','')}</div>
        <div>{tags_html}</div>
    </div>
    """, unsafe_allow_html=True)

    # ── 종합 점수 (정량화) ──
    if ov_score is not None:
        ovc = overall_color(ov_score)
        st.markdown(f"""
        <div style="background:white;border:2px solid {ovc};border-radius:12px;
                    padding:1.3rem 1.8rem;margin-bottom:1.5rem;
                    display:flex;align-items:center;justify-content:space-between;">
            <div>
                <div style="font-size:0.6rem;font-weight:700;letter-spacing:3px;
                            text-transform:uppercase;color:{ovc};">
                    Overall Score · 종합 점수
                </div>
                <div style="font-size:0.72rem;color:#7A7268;margin-top:0.4rem;line-height:1.6;">
                    세부 점수 가중합 · 역량 35% · 조직적합 30% · 리더십준비 15% · 저위험 20%
                </div>
            </div>
            <div style="text-align:right;line-height:1;">
                <span style="font-family:'DM Serif Display',serif;font-size:3rem;
                             font-style:italic;color:{ovc};">{ov_score}<span style="font-size:1rem;color:#B0A898;">/100</span></span>
                <div style="font-size:1.05rem;font-weight:800;color:{ovc};margin-top:0.2rem;">{ov_grade}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ── 프로필 & 분석 자료 커버리지 ──
    prof = R.get("profile", {})
    cov  = R.get("data_coverage", {})
    if prof or cov:
        prof_bits = []
        for lbl, key in [("전공","major"),("대학","university"),("학력","education_level"),("출신지역","region")]:
            val = prof.get(key)
            if val and val != "자료 미제공":
                prof_bits.append(f'<b style="color:#3D3830;">{lbl}</b> {val}')
        prof_html = " &nbsp;·&nbsp; ".join(prof_bits) if prof_bits else '<span style="color:#B0A898;">프로필 자료 미확인</span>'
        conf = cov.get("confidence","")
        conf_color = {"HIGH":"#2D6A4F","MEDIUM":"#8B6914","LOW":"#8B2635"}.get(conf, "#7A7268")
        mats_html = "".join(
            f'<span style="display:inline-block;background:#F2EEE6;border:1px solid #E2DDD4;'
            f'border-radius:4px;padding:1px 7px;margin:1px;font-size:0.66rem;color:#3D3830;">{m}</span>'
            for m in cov.get("materials", [])
        )
        conf_badge = (f'<span style="display:inline-block;background:{conf_color};color:white;'
                      f'border-radius:4px;padding:1px 8px;font-size:0.66rem;font-weight:700;">근거 신뢰도 {conf or "—"}</span>') if cov else ""
        st.markdown(f"""
        <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;
                    padding:0.9rem 1.3rem;margin-bottom:1.5rem;">
            <div style="font-size:0.78rem;color:#3D3830;line-height:1.8;">{prof_html}</div>
            <div style="margin-top:0.5rem;display:flex;align-items:center;gap:0.4rem;flex-wrap:wrap;">
                {conf_badge}{mats_html}
            </div>
            <div style="font-size:0.7rem;color:#7A7268;margin-top:0.4rem;">{cov.get('note','')}</div>
        </div>
        """, unsafe_allow_html=True)

    # ── 리밸런싱 판정 배너 ──
    rv = R.get("rebalancing_verdict", {})
    if rv:
        decision = rv.get("decision", "—")
        confidence = rv.get("confidence", "—")
        dec_style = {
            "KEEP":    ("#2D6A4F", "#EAF4EE", "✓ KEEP — 핵심 인재, 유지 권장"),
            "DEVELOP": ("#2B3D5C", "#E8EEF5", "↗ DEVELOP — 육성 대상"),
            "WATCH":   ("#8B6914", "#FBF3E0", "◷ WATCH — 관찰 필요"),
            "MISFIT":  ("#8B2635", "#FAEAEC", "✕ MISFIT — 조직 방향성 부적합"),
        }
        color, bg, label = dec_style.get(decision, ("#7A7268", "#EDE8E0", decision))
        st.markdown(f"""
        <div style="background:{bg};border:2px solid {color};border-radius:10px;
                    padding:1.3rem 1.8rem;margin-bottom:1.5rem;
                    display:flex;align-items:center;justify-content:space-between;">
            <div>
                <div style="font-size:0.6rem;font-weight:700;letter-spacing:3px;
                            text-transform:uppercase;color:{color};margin-bottom:0.3rem;">
                    Rebalancing Verdict
                </div>
                <div style="font-size:1.1rem;font-weight:800;color:{color};">{label}</div>
            </div>
            <div style="text-align:right;">
                <div style="font-size:0.6rem;color:#7A7268;letter-spacing:1px;">신뢰도</div>
                <div style="font-size:0.95rem;font-weight:700;color:{color};">{confidence}</div>
            </div>
        </div>
        <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;
                    padding:1rem 1.3rem;margin-bottom:1.5rem;">
            <div style="font-size:0.83rem;color:#3D3830;line-height:1.75;">
                {rv.get('rationale','')}
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ── 조직 적합도 & 리더십 준비도 ──
    ofit = R.get("org_fit", {})
    lead = R.get("leadership_readiness", {})
    if ofit or lead:
        st.markdown("""
        <div class="section-header">
            <span class="section-num">02</span>
            <span class="section-title">조직 적합도 & 리더십 준비도</span>
            <div class="section-rule"></div>
        </div>
        """, unsafe_allow_html=True)

        of_col, lr_col = st.columns(2)
        with of_col:
            of_score = ofit.get("score", 0)
            of_color = "#2D6A4F" if of_score >= 70 else ("#8B6914" if of_score >= 50 else "#8B2635")
            st.markdown(f"""
            <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;
                        padding:1.4rem 1.6rem;border-left:3px solid {of_color};">
                <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:0.6rem;">
                    <span style="font-size:0.65rem;font-weight:700;letter-spacing:2px;
                                 text-transform:uppercase;color:{of_color};">🎯 조직 적합도</span>
                    <span style="font-family:'DM Serif Display',serif;font-size:1.8rem;
                                 font-style:italic;color:{of_color};">{of_score}<span style="font-size:0.8rem;color:#B0A898;">/100</span></span>
                </div>
                <div style="background:#E2DDD4;border-radius:999px;height:4px;margin-bottom:0.9rem;overflow:hidden;">
                    <div style="width:{of_score}%;height:100%;background:{of_color};border-radius:999px;"></div>
                </div>
                <div style="font-size:0.78rem;color:#3D3830;line-height:1.7;margin-bottom:0.6rem;">{ofit.get('summary','')}</div>
                <div style="border-top:1px solid #E2DDD4;padding-top:0.6rem;font-size:0.73rem;color:#7A7268;line-height:1.6;">
                    <b style="color:#3D3830;">인재상</b> {ofit.get('talent_match','')}<br>
                    <b style="color:#3D3830;">문화</b> {ofit.get('culture_fit','')}<br>
                    <b style="color:#3D3830;">방향성</b> {ofit.get('direction_alignment','')}
                </div>
            </div>
            """, unsafe_allow_html=True)
        with lr_col:
            lr_score = lead.get("score", 0)
            lr_color = "#2D6A4F" if lr_score >= 80 else ("#2B3D5C" if lr_score >= 60 else "#8B2635")
            st.markdown(f"""
            <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;
                        padding:1.4rem 1.6rem;border-left:3px solid {lr_color};">
                <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:0.6rem;">
                    <span style="font-size:0.65rem;font-weight:700;letter-spacing:2px;
                                 text-transform:uppercase;color:{lr_color};">👑 리더십 준비도</span>
                    <span style="font-family:'DM Serif Display',serif;font-size:1.8rem;
                                 font-style:italic;color:{lr_color};">{lr_score}<span style="font-size:0.8rem;color:#B0A898;">/100</span></span>
                </div>
                <div style="background:#E2DDD4;border-radius:999px;height:4px;margin-bottom:0.9rem;overflow:hidden;">
                    <div style="width:{lr_score}%;height:100%;background:{lr_color};border-radius:999px;"></div>
                </div>
                <div style="display:inline-block;background:{lr_color};color:white;border-radius:4px;
                            padding:0.2rem 0.7rem;font-size:0.72rem;font-weight:700;margin-bottom:0.7rem;">
                    {lead.get('level','—')}
                </div>
                <div style="font-size:0.78rem;color:#3D3830;line-height:1.7;margin-bottom:0.6rem;">{lead.get('rationale','')}</div>
                <div style="border-top:1px solid #E2DDD4;padding-top:0.6rem;font-size:0.75rem;color:#3D3830;line-height:1.6;">
                    <b style="color:{lr_color};">리더 부여 결론</b><br>{lead.get('recommendation','')}
                </div>
            </div>
            """, unsafe_allow_html=True)

    # ── 회사 향후 방향성 적합도 ──
    df = R.get("direction_fit", {})
    if df and isinstance(df, dict) and (df.get("summary") or df.get("score")):
        df_score = df.get("score", 0) or 0
        df_color = overall_color(df_score)
        st.markdown(f"""
        <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;
                    padding:1.4rem 1.6rem;margin:1.3rem 0 1.5rem;border-left:3px solid {df_color};">
            <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:0.6rem;">
                <span style="font-size:0.65rem;font-weight:700;letter-spacing:2px;
                             text-transform:uppercase;color:{df_color};">🚀 회사 향후 방향성 적합도</span>
                <span style="font-family:'DM Serif Display',serif;font-size:1.6rem;font-style:italic;color:{df_color};">{df_score}<span style="font-size:0.75rem;color:#B0A898;">/100</span></span>
            </div>
            <div style="background:#E2DDD4;border-radius:999px;height:4px;margin-bottom:0.8rem;overflow:hidden;">
                <div style="width:{df_score}%;height:100%;background:{df_color};border-radius:999px;"></div>
            </div>
            <div style="font-size:0.8rem;color:#3D3830;line-height:1.7;">{df.get('summary','')}</div>
        </div>
        """, unsafe_allow_html=True)

    # ── 역량 차원 ──
    st.markdown("""
    <div class="section-header">
        <span class="section-num">03</span>
        <span class="section-title">역량 차원 분석</span>
        <div class="section-rule"></div>
    </div>
    """, unsafe_allow_html=True)

    dims = R.get("dimensions", {})
    d_col1, d_col2 = st.columns(2)
    for idx, (key, info) in enumerate(dims.items()):
        score  = info.get("score", 0)
        grade  = info.get("grade", "-")
        gcolor = grade_color(grade)
        ev_html = "".join(
            f'<div class="evidence-item"><div class="evidence-dot"></div><span>{e}</span></div>'
            for e in info.get("evidence", [])
        )
        subs = info.get("sub_scores", {})
        sub_html = ""
        if isinstance(subs, dict) and subs:
            chips = "".join(
                f'<span style="display:inline-block;background:#F2EEE6;border:1px solid #E2DDD4;'
                f'border-radius:4px;padding:2px 8px;margin:2px;font-size:0.68rem;color:#3D3830;">'
                f'{k} <b style="color:{gcolor};">{v}</b></span>'
                for k, v in subs.items()
            )
            sub_html = f'<div style="margin:0.3rem 0 0.5rem;">{chips}</div>'
        with (d_col1 if idx % 2 == 0 else d_col2):
            st.markdown(f"""
            <div class="dim-card">
                <div class="dim-header">
                    <div class="dim-icon-title">
                        <div class="dim-icon">{dim_icon(key)}</div>
                        <div>
                            <div class="dim-name">{dim_label(key)}</div>
                            <div class="dim-sub">{dim_sublabel(key)}</div>
                        </div>
                    </div>
                    <div class="dim-score-block">
                        <div class="dim-score">{score}<span>/100</span></div>
                        <div class="dim-grade" style="color:{gcolor};">{grade}</div>
                    </div>
                </div>
                <div class="progress-track">
                    <div class="progress-fill" style="width:{score}%;"></div>
                </div>
                <div class="dim-summary">{info.get('summary','')}</div>
                {sub_html}
                <div class="evidence-list">{ev_html}</div>
            </div>
            """, unsafe_allow_html=True)

    # ── 학력-성과 정합성 & 커리어 트랙 ──
    cp = R.get("credential_performance", {})
    ct = R.get("career_track", {})
    if cp or ct:
        st.markdown("""
        <div class="section-header" style="margin-top:2.5rem;">
            <span class="section-num">05</span>
            <span class="section-title">학력-성과 정합성 & 커리어 트랙</span>
            <div class="section-rule"></div>
        </div>
        """, unsafe_allow_html=True)

        cp_col, ct_col = st.columns(2)
        with cp_col:
            align = cp.get("alignment", "—")
            align_color = {"일치":"#2D6A4F","학력우위":"#8B2635","성과우위":"#2B3D5C"}.get(align, "#7A7268")
            st.markdown(f"""
            <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;
                        padding:1.4rem 1.6rem;border-left:3px solid {align_color};height:100%;">
                <div style="font-size:0.65rem;font-weight:700;letter-spacing:2px;
                            text-transform:uppercase;color:{align_color};margin-bottom:0.6rem;">
                    🎓 학력-성과 정합성
                </div>
                <div style="display:inline-block;background:{align_color};color:white;border-radius:4px;
                            padding:0.2rem 0.7rem;font-size:0.72rem;font-weight:700;margin-bottom:0.7rem;">
                    {align}
                </div>
                <div style="font-size:0.72rem;color:#7A7268;margin-bottom:0.5rem;">
                    학력 수준: {cp.get('education_level','자료 미제공')}
                </div>
                <div style="font-size:0.8rem;color:#3D3830;line-height:1.7;">{cp.get('summary','')}</div>
            </div>
            """, unsafe_allow_html=True)
        with ct_col:
            st.markdown(f"""
            <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;
                        padding:1.4rem 1.6rem;border-left:3px solid #B8924A;height:100%;">
                <div style="font-size:0.65rem;font-weight:700;letter-spacing:2px;
                            text-transform:uppercase;color:#B8924A;margin-bottom:0.6rem;">
                    🧭 커리어 트랙
                </div>
                <div style="font-size:0.72rem;color:#7A7268;margin-bottom:0.3rem;">
                    현재: {ct.get('current_position','—')}
                </div>
                <div style="display:inline-block;background:#B8924A;color:white;border-radius:4px;
                            padding:0.2rem 0.7rem;font-size:0.72rem;font-weight:700;margin-bottom:0.7rem;">
                    {ct.get('recommended_track','—')}
                </div>
                <div style="font-size:0.8rem;color:#3D3830;line-height:1.7;">{ct.get('summary','')}</div>
            </div>
            """, unsafe_allow_html=True)

    # ── 채용 키워드 ──
    st.markdown("""
    <div class="section-header" style="margin-top:2.5rem;">
        <span class="section-num">06</span>
        <span class="section-title">채용 핵심 키워드 Top 3</span>
        <div class="section-rule"></div>
    </div>
    """, unsafe_allow_html=True)

    rank_cls = ["rank-gold","rank-silver","rank-bronze"]
    rank_lbl = ["1st","2nd","3rd"]
    for kw in R.get("hiring_keywords", []):
        r  = kw.get("rank", 1) - 1
        rc = rank_cls[r] if r < 3 else "rank-bronze"
        rl = rank_lbl[r] if r < 3 else f"{r+1}th"
        st.markdown(f"""
        <div class="kw-card">
            <div class="kw-rank-col">
                <div class="kw-rank-num {rc}">{r+1}</div>
                <span class="kw-rank-label">{rl}</span>
            </div>
            <div>
                <div class="kw-title">{kw.get('keyword','')}</div>
                <div class="kw-why">{kw.get('why','')}</div>
                <div class="kw-how">
                    <div class="kw-how-label">확인 방법</div>
                    <div class="kw-how-text">{kw.get('how_to_check','')}</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ── 번아웃 & 이직 가능성 ──
    st.markdown("""
    <div class="section-header" style="margin-top:2.5rem;">
        <span class="section-num">07</span>
        <span class="section-title">번아웃 위험도 & 이직 가능성</span>
        <div class="section-rule"></div>
    </div>
    """, unsafe_allow_html=True)

    burnout  = R.get("burnout_risk", {})
    turnover = R.get("turnover_risk", {})

    def risk_color(level):
        return {"LOW":"#2D6A4F","MEDIUM":"#8B6914","HIGH":"#8B2635","CRITICAL":"#5C0011"}.get(level,"#7A7268")
    def risk_label(level):
        return {"LOW":"낮음","MEDIUM":"주의","HIGH":"높음","CRITICAL":"심각"}.get(level, level)
    def risk_bar_color(level):
        return {"LOW":"#2D6A4F","MEDIUM":"#D4AF72","HIGH":"#C0392B","CRITICAL":"#7B0000"}.get(level,"#B0A898")

    b_level  = burnout.get("level","MEDIUM")
    b_score  = burnout.get("score", 50)
    t_level  = turnover.get("level","LOW")
    t_score  = turnover.get("score", 30)

    b_col, t_col = st.columns(2)

    with b_col:
        bcolor = risk_color(b_level)
        st.markdown(f"""
        <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;
                    padding:1.4rem 1.6rem;border-left:3px solid {bcolor};">
            <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:0.8rem;">
                <div style="font-size:0.65rem;font-weight:700;letter-spacing:3px;
                            text-transform:uppercase;color:{bcolor};">🔥 번아웃 위험도</div>
                <span style="background:{bcolor};color:white;border-radius:4px;
                             padding:0.2rem 0.7rem;font-size:0.7rem;font-weight:700;">
                    {risk_label(b_level)}
                </span>
            </div>
            <div style="background:#E2DDD4;border-radius:999px;height:5px;margin-bottom:1rem;overflow:hidden;">
                <div style="width:{b_score}%;height:100%;background:{risk_bar_color(b_level)};border-radius:999px;"></div>
            </div>
            <div style="font-size:0.8rem;color:#3D3830;line-height:1.75;margin-bottom:0.8rem;">
                {burnout.get('summary','자료 부족으로 분석 불가')}
            </div>
            <div style="border-top:1px solid #E2DDD4;padding-top:0.8rem;">
                <div style="font-size:0.68rem;color:#B0A898;font-weight:600;margin-bottom:0.4rem;letter-spacing:1px;">세부 분석</div>
                <div style="font-size:0.75rem;color:#7A7268;margin-bottom:0.3rem;">
                    <b style="color:#3D3830;">정서적 고갈</b> — {burnout.get('emotional_exhaustion','')}
                </div>
                <div style="font-size:0.75rem;color:#7A7268;margin-bottom:0.3rem;">
                    <b style="color:#3D3830;">비인격화</b> — {burnout.get('depersonalization','')}
                </div>
                <div style="font-size:0.75rem;color:#7A7268;">
                    <b style="color:#3D3830;">성취감</b> — {burnout.get('personal_accomplishment','')}
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    with t_col:
        tcolor = risk_color(t_level)
        st.markdown(f"""
        <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;
                    padding:1.4rem 1.6rem;border-left:3px solid {tcolor};">
            <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:0.8rem;">
                <div style="font-size:0.65rem;font-weight:700;letter-spacing:3px;
                            text-transform:uppercase;color:{tcolor};">🚪 이직 가능성</div>
                <span style="background:{tcolor};color:white;border-radius:4px;
                             padding:0.2rem 0.7rem;font-size:0.7rem;font-weight:700;">
                    {risk_label(t_level)}
                </span>
            </div>
            <div style="background:#E2DDD4;border-radius:999px;height:5px;margin-bottom:1rem;overflow:hidden;">
                <div style="width:{t_score}%;height:100%;background:{risk_bar_color(t_level)};border-radius:999px;"></div>
            </div>
            <div style="font-size:0.8rem;color:#3D3830;line-height:1.75;margin-bottom:0.8rem;">
                {turnover.get('summary','자료 부족으로 분석 불가')}
            </div>
            <div style="border-top:1px solid #E2DDD4;padding-top:0.8rem;">
                <div style="font-size:0.68rem;color:#B0A898;font-weight:600;margin-bottom:0.4rem;letter-spacing:1px;">세부 분석</div>
                <div style="font-size:0.75rem;color:#7A7268;margin-bottom:0.3rem;">
                    <b style="color:#3D3830;">이탈 요인</b> — {turnover.get('push_factors','')}
                </div>
                <div style="font-size:0.75rem;color:#7A7268;margin-bottom:0.3rem;">
                    <b style="color:#3D3830;">외부 유인</b> — {turnover.get('pull_factors','')}
                </div>
                <div style="font-size:0.75rem;color:#7A7268;">
                    <b style="color:#3D3830;">재직 신호</b> — {turnover.get('retention_signals','')}
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ── Derailer & Development ──
    st.markdown("""
    <div class="section-header" style="margin-top:2.5rem;">
        <span class="section-num">08</span>
        <span class="section-title">리스크 & 개발 제언</span>
        <div class="section-rule"></div>
    </div>
    """, unsafe_allow_html=True)

    dd1, dd2 = st.columns(2)
    with dd1:
        st.markdown(f"""
        <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;
                    padding:1.4rem 1.6rem;border-left:3px solid #8B2635;">
            <div style="font-size:0.65rem;font-weight:700;letter-spacing:3px;
                        text-transform:uppercase;color:#8B2635;margin-bottom:0.8rem;">
                ⚠ Derailer · 잠재적 위험 요인
            </div>
            <div style="font-size:0.85rem;color:#3D3830;line-height:1.85;">
                {R.get('derailer','자료 부족으로 분석 불가')}
            </div>
        </div>
        """, unsafe_allow_html=True)
    with dd2:
        st.markdown(f"""
        <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;
                    padding:1.4rem 1.6rem;border-left:3px solid #2D6A4F;">
            <div style="font-size:0.65rem;font-weight:700;letter-spacing:3px;
                        text-transform:uppercase;color:#2D6A4F;margin-bottom:0.8rem;">
                ◆ Development · 성과 극대화 조건
            </div>
            <div style="font-size:0.85rem;color:#3D3830;line-height:1.85;">
                {R.get('development_suggestion','자료 부족으로 분석 불가')}
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ── SNS 전용 분석 (대외비) ──
    sns = R.get("sns_analysis", {})
    if sns and sns.get("available"):
        sns_score = sns.get("score", 0) or 0
        sns_color = overall_color(sns_score)
        st.markdown("""
        <div class="section-header" style="margin-top:2.5rem;">
            <span class="section-num">✦</span>
            <span class="section-title">SNS 기반 성향·문화 적합성 &nbsp;<span style="font-size:0.6rem;background:#8B2635;color:white;padding:1px 7px;border-radius:4px;vertical-align:middle;letter-spacing:1px;">대외비</span></span>
            <div class="section-rule"></div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown(f"""
        <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;
                    padding:1.4rem 1.6rem;margin-bottom:1.5rem;border-left:3px solid {sns_color};">
            <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:0.6rem;">
                <span style="font-size:0.65rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:{sns_color};">📱 SNS 적합도</span>
                <span style="font-family:'DM Serif Display',serif;font-size:1.6rem;font-style:italic;color:{sns_color};">{sns_score}<span style="font-size:0.75rem;color:#B0A898;">/100</span></span>
            </div>
            <div style="font-size:0.78rem;color:#3D3830;line-height:1.7;margin-bottom:0.5rem;"><b>성향</b> &nbsp;{sns.get('personality','')}</div>
            <div style="font-size:0.78rem;color:#3D3830;line-height:1.7;"><b>조직문화 적합</b> &nbsp;{sns.get('culture_fit','')}</div>
        </div>
        """, unsafe_allow_html=True)

    # ── Overall Insight ──
    st.markdown("""
    <div class="section-header" style="margin-top:2.5rem;">
        <span class="section-num">09</span>
        <span class="section-title">종합 인사이트 & 리밸런싱 권고</span>
        <div class="section-rule"></div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown(f"""
    <div class="insight-box">
        <div class="insight-label">◈ Executive Assessment Summary</div>
        <div class="insight-text">{R.get('overall_insight','')}</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div style="text-align:center;padding:2.5rem 0 1rem;font-size:0.65rem;
         letter-spacing:3px;text-transform:uppercase;color:#B0A898;">
        Assessment Complete &nbsp;·&nbsp; M.I.Tech Talent Intelligence &nbsp;·&nbsp; Confidential
    </div>
    """, unsafe_allow_html=True)


# ─── 사이드바 아카이브 ──────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="padding:0.5rem 0 1rem;">
        <div style="font-family:'DM Serif Display',serif;font-size:1.2rem;
                    color:#1A1714;font-style:italic;margin-bottom:0.3rem;">
            Archive
        </div>
        <div style="font-size:0.65rem;letter-spacing:2px;text-transform:uppercase;
                    color:#B0A898;">분석 기록 조회</div>
        <div style="height:1px;background:#D4CEC4;margin-top:0.8rem;"></div>
    </div>
    """, unsafe_allow_html=True)

    archive = load_archive()

    if not archive:
        st.markdown('<p style="font-size:0.8rem;color:#B0A898;text-align:center;padding:2rem 0;">저장된 분석 기록이 없습니다.</p>', unsafe_allow_html=True)
    else:
        # 전체 내보내기
        export_json = json.dumps(archive, ensure_ascii=False, indent=2)
        st.download_button(
            label="⬇ 전체 기록 내보내기 (JSON)",
            data=export_json.encode("utf-8"),
            file_name=f"talent_archive_{datetime.now().strftime('%Y%m%d')}.json",
            mime="application/json",
            use_container_width=True
        )
        st.markdown('<div style="height:0.5rem;"></div>', unsafe_allow_html=True)

        for i, rec in enumerate(archive):
            name    = rec.get("candidate_name", "이름 없음")
            dept    = rec.get("dept", "")
            saved   = rec.get("saved_at", "")
            rec_id  = rec.get("id", None)
            summary = rec.get("result", {}).get("candidate_summary", "")[:40]
            tags    = rec.get("result", {}).get("personality_tags", [])[:2]

            # 카드 클릭 = 해당 기록 조회
            with st.expander(f"**{name}** · {saved}", expanded=False):
                if dept:
                    st.caption(dept)
                st.markdown(f'<p style="font-size:0.75rem;color:#7A7268;line-height:1.6;">{summary}...</p>',
                            unsafe_allow_html=True)
                for t in tags:
                    st.markdown(f'<span style="display:inline-block;border:1px solid #D4AF72;color:#B8924A;'
                                f'border-radius:3px;padding:1px 7px;font-size:0.68rem;margin:1px;">{t}</span>',
                                unsafe_allow_html=True)
                col_v, col_d = st.columns(2)
                with col_v:
                    if st.button("조회", key=f"view_{i}", use_container_width=True):
                        st.session_state["archive_view"] = i
                        st.session_state["show_archive"] = True
                        st.rerun()
                with col_d:
                    if st.button("삭제", key=f"del_{i}", use_container_width=True):
                        delete_from_archive(rec_id, i)
                        if st.session_state.get("archive_view") == i:
                            st.session_state["show_archive"] = False
                        st.rerun()

# ─── UI ─────────────────────────────────────────────────────────────────────

# Masthead
st.markdown("""
<div class="masthead">
    <div>
        <div class="masthead-brand">Talent <em>Intelligence</em></div>
    </div>
    <div class="masthead-meta">
        M.I.TECH · P&C TEAM<br>
        인재 심층 분석 플랫폼<br>
        INTERNAL USE ONLY
    </div>
</div>
<div class="masthead-rule"></div>
<div class="masthead-thin"></div>
""", unsafe_allow_html=True)

# ── API Key: Streamlit Secrets에서 자동 로드 ──
try:
    api_key = st.secrets["ANTHROPIC_API_KEY"]
except Exception:
    st.error("⚠️ 서버 설정 오류입니다. 관리자에게 문의해주세요.")
    st.stop()

# 공통 인재상 (두 탭에서 공유)
company_standard = """1) 성장지향: 목표 의식이 뚜렷하며, 조직과 개인의 동반 성장을 위해 노력하는 분
2) 상호존중: 동료 간의 상호 존중과 팀워크의 가치를 소중히 여기는 분
3) 혁신과 도전: 지속적인 학습과 도전을 통해 끊임없이 혁신을 추구하는 분"""

core_culture = """1) 개방적 소통 탁월성 (Open communication excellence)
2) 몰입 기반 실행력 (Commitment-driven execution)
3) 성과 기반 인정 체계 (Performance-based recognition)
4) 협업 시너지 (Collaborative synergy)
5) 혁신 리더십 (Innovation leadership)"""

# 회사 향후 방향성 (방향성 적합도 분석에 사용 · 화면에서 수정 가능)
company_direction = """※ 회사의 향후 전략 방향을 여기에 입력하세요. (예시)
- 비혈관 스텐트 글로벌 시장 확대 및 해외 인허가 가속
- R&D 기반 신제품 파이프라인 강화
- 데이터·디지털 기반 품질·생산 고도화
- 성과 중심·협업 중심의 조직문화 정착"""

# ── 8대 표준 분석자료 ──
STANDARD_MATERIALS = [
    "이력서", "포트폴리오", "다면평가 결과", "기안서",
    "MBTI 결과", "인적성 검사", "SNS", "기타자료",
]
# 항목별 파일명 키워드 (위에서부터 우선 매칭)
MATERIAL_KEYWORDS = {
    "이력서": ["이력서", "resume", "cv", "경력기술", "자기소개", "자소서", "프로필", "profile"],
    "포트폴리오": ["포트폴리오", "portfolio", "작품", "성과물"],
    "다면평가 결과": ["다면평가", "다면", "360", "동료평가", "peer", "리뷰", "review"],
    "기안서": ["기안", "품의", "제안서", "보고서", "기획서", "draft", "proposal"],
    "MBTI 결과": ["mbti", "엠비티아이", "성격유형", "16personalities"],
    "인적성 검사": ["인적성", "적성검사", "인성검사", "적성", "인성", "aptitude", "assessment"],
    "SNS": ["sns", "인스타", "instagram", "facebook", "페북", "블로그", "blog",
            "linkedin", "링크드인", "트위터", "twitter", "유튜브", "youtube", "틱톡", "tiktok"],
}

def classify_material(filename):
    """파일명을 8대 표준 자료 항목 중 하나로 분류."""
    low = (filename or "").lower()
    for cat, kws in MATERIAL_KEYWORDS.items():
        for kw in kws:
            if kw.lower() in low:
                return cat
    return "기타자료"

def coverage_from_filenames(filenames):
    """파일명 리스트 → 충족된 표준 자료 항목 set."""
    covered = set()
    for fn in filenames or []:
        covered.add(classify_material(fn))
    return covered

def material_checklist_html(covered, title="8대 표준 분석자료 충족 현황"):
    """충족 현황을 8개 칩으로 표시하는 HTML."""
    covered = set(covered or [])
    chips = []
    for m in STANDARD_MATERIALS:
        on = m in covered
        chips.append(
            f'<span style="display:inline-flex;align-items:center;gap:5px;'
            f'background:{"#E6F2EA" if on else "#F2EEE6"};'
            f'border:1px solid {"#9CCBB0" if on else "#E2DDD4"};'
            f'border-radius:6px;padding:3px 10px;margin:3px;font-size:0.76rem;'
            f'color:{"#1E5C3A" if on else "#B0A898"};font-weight:{700 if on else 500};">'
            f'{"✅" if on else "⚪"} {m}</span>'
        )
    cnt = len([m for m in STANDARD_MATERIALS if m in covered])
    return (
        f'<div style="background:white;border:1px solid #D4CEC4;border-radius:8px;padding:0.8rem 1rem;margin:0.4rem 0;">'
        f'<div style="font-size:0.72rem;font-weight:700;letter-spacing:1px;color:#B8924A;'
        f'text-transform:uppercase;margin-bottom:0.5rem;">📋 {title} · {cnt}/8</div>'
        f'<div>{"".join(chips)}</div></div>'
    )

def material_chips_inline(covered):
    """인원별 컴팩트 충족 표시 (8개 작은 칩)."""
    covered = set(covered or [])
    parts = []
    for m in STANDARD_MATERIALS:
        on = m in covered
        parts.append(
            f'<span style="font-size:0.66rem;padding:1px 6px;border-radius:4px;margin:1px;display:inline-block;'
            f'background:{"#E6F2EA" if on else "#F4F1EB"};color:{"#1E5C3A" if on else "#C0BCB4"};'
            f'border:1px solid {"#9CCBB0" if on else "#E8E3DA"};">{("✓ " if on else "") + m}</span>'
        )
    cnt = len([m for m in STANDARD_MATERIALS if m in covered])
    return (f'<div style="margin:3px 0 7px;">'
            f'<span style="font-size:0.68rem;color:#7A7268;font-weight:700;margin-right:4px;">📋 충족 {cnt}/8</span>'
            f'{"".join(parts)}</div>')

# ── 탭 ──
tab_single, tab_bulk, tab_org = st.tabs(["👤  개인 분석", "👥  대량 분석", "🏢  조직도"])


# ════════════════════════════════════════════════════════
#  TAB 1 — 개인 분석 (기존 기능)
# ════════════════════════════════════════════════════════
with tab_single:
    # ── 대상자 기본 정보 ──
    st.markdown("""
    <div class="section-header">
        <span class="section-num">01</span>
        <span class="section-title">대상자 기본 정보</span>
        <div class="section-rule"></div>
    </div>
    """, unsafe_allow_html=True)
    # ── 검사 대상자 선택 (조직 명부 / 직접 입력) ──
    _org_s = load_org_data()
    _done_names = {r.get("candidate_name", "") for r in load_archive() if r.get("candidate_name")}

    def _flatten_org(org):
        out = {}
        for bonbu, bd in org.items():
            units = {}
            bdir = bd.get("_직속", {})
            dp = bdir.get("_직속", []) if isinstance(bdir, dict) else []
            if dp:
                units["(본부 직속)"] = dp
            for tname, tdict in bd.items():
                if tname == "_직속":
                    continue
                td = tdict.get("_직속", [])
                if td:
                    units[tname] = td
                for pname, plist in tdict.items():
                    if pname == "_직속":
                        continue
                    units[f"{tname} · {pname}"] = plist
            if units:
                out[bonbu] = units
        return out

    s_mode = st.radio("대상자 선택 방식", ["조직 명부에서 선택", "직접 입력"], horizontal=True, key="s_mode")
    candidate_name = ""
    candidate_dept = ""

    if s_mode == "조직 명부에서 선택" and _org_s:
        _flat = _flatten_org(_org_s)
        _all = [p["name"] for u in _flat.values() for lst in u.values() for p in lst]
        _tot = len(_all)
        _dn = sum(1 for n in _all if n in _done_names)
        st.markdown(
            f'<div style="display:flex;gap:1.2rem;align-items:center;flex-wrap:wrap;margin-bottom:0.7rem;font-size:0.82rem;color:#7A7268;">'
            f'<span>전체 <b style="color:#1A1714;">{_tot}</b>명</span>'
            f'<span>✅ 등록 <b style="color:#2D6A4F;">{_dn}</b></span>'
            f'<span>⚪ 미등록 <b style="color:#B0A898;">{_tot - _dn}</b></span></div>',
            unsafe_allow_html=True)
        pc1, pc2, pc3 = st.columns([1, 1.4, 1.5])
        with pc1:
            s_bonbu = st.selectbox("본부", list(_flat.keys()), key="s_bonbu")
        _units = _flat.get(s_bonbu, {})
        with pc2:
            s_unit = st.selectbox("팀 · 파트", list(_units.keys()), key="s_unit") if _units else None
        _members = _units.get(s_unit, []) if s_unit else []
        _optmap, _opts = {}, []
        for p in _members:
            tag = "✅" if p["name"] in _done_names else "⚪"
            lab = f'{tag} {p["name"]} · {p.get("pos", "")}'
            _opts.append(lab)
            _optmap[lab] = p
        with pc3:
            s_pick = st.selectbox("인원 (✅ 등록 / ⚪ 미등록)", _opts, key="s_pick") if _opts else None
        if s_pick:
            _ps = _optmap[s_pick]
            candidate_name = _ps["name"]
            candidate_dept = f"{s_bonbu} / {s_unit}" if s_unit else s_bonbu
            _reg = candidate_name in _done_names
            _badge = ('<span style="color:#2D6A4F;font-weight:700;">✅ 이미 분석 등록됨 — 다시 검사하면 최신 결과로 갱신됩니다</span>'
                      if _reg else
                      '<span style="color:#8B6914;font-weight:700;">⚪ 미등록 — 자료를 올려 첫 검사를 진행하세요</span>')
            st.markdown(
                f'<div style="background:#FFFFFF;border:1px solid #D4CEC4;border-left:3px solid #B8924A;'
                f'border-radius:8px;padding:0.7rem 1rem;margin:0.3rem 0 0.2rem;font-size:0.85rem;">'
                f'선택: <b style="color:#1A1714;">{candidate_name}</b> '
                f'<span style="color:#7A7268;">· {candidate_dept}</span><br>{_badge}</div>',
                unsafe_allow_html=True)
        with st.expander("📋 조직 명부 분석 등록 현황 (전체 보기)"):
            for bonbu, units in _flat.items():
                bt = sum(len(l) for l in units.values())
                bd_ = sum(1 for l in units.values() for p in l if p["name"] in _done_names)
                st.markdown(f"**{bonbu}** · 등록 {bd_}/{bt}")
                lines = []
                for u, lst in units.items():
                    chips = " ".join(("✅" if p["name"] in _done_names else "⚪") + p["name"] for p in lst)
                    lines.append(f"- {u} : {chips}")
                st.markdown("\n".join(lines))
    else:
        c1, c2 = st.columns(2)
        with c1:
            candidate_name = st.text_input("성명", value=st.session_state.get("org_prefill_name", ""), placeholder="홍길동", key="s_name")
        with c2:
            candidate_dept = st.text_input("소속 부서", placeholder="Sales & Marketing Division", key="s_dept")
    company_standard_s = st.text_area("회사 인재상", value=company_standard, height=110, key="s_std")
    core_culture_s     = st.text_area("회사 5대 핵심문화 축", value=core_culture, height=140, key="s_culture")
    company_direction_s = st.text_area("회사 향후 방향성 (방향성 적합도 분석에 사용)", value=company_direction, height=120, key="s_direction")

    # ── 파일 업로드 ──
    st.markdown("""
    <div class="section-header">
        <span class="section-num">02</span>
        <span class="section-title">자료 업로드</span>
        <div class="section-rule"></div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown(("""
    <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;padding:1.2rem 1.5rem;margin-bottom:1rem;">
        <p style="font-size:0.8rem;font-weight:600;color:#1A1714;margin-bottom:0.6rem;">📎 분석에 활용되는 8대 표준 자료 — 아래 항목을 한꺼번에 선택해서 업로드하세요</p>
        <div style="display:flex;flex-wrap:wrap;gap:0.4rem 1.6rem;">"""
        + "".join(f'<span style="font-size:0.75rem;color:#7A7268;">✦ {i+1}. {m}</span>' for i, m in enumerate(STANDARD_MATERIALS))
        + """</div>
        <p style="font-size:0.7rem;color:#B0A898;margin-top:0.8rem;margin-bottom:0;">PDF · DOCX · JPG · PNG · TXT 지원 · 파일명에 자료 종류(예: 홍길동_이력서, 홍길동_MBTI)를 넣으면 충족 항목이 자동 인식됩니다</p>
    </div>
    """), unsafe_allow_html=True)
    uploaded_files = st.file_uploader(
        "파일을 여기에 끌어다 놓거나 클릭해서 선택하세요 (여러 파일 동시 선택 가능)",
        type=["pdf","docx","jpg","jpeg","png","webp","txt","md"],
        accept_multiple_files=True, label_visibility="visible", key="s_uploader"
    )
    file_data = {}
    if uploaded_files:
        st.markdown(f'<p style="font-size:0.78rem;color:#2D6A4F;margin:0.5rem 0;">✅ {len(uploaded_files)}개 파일 업로드 완료</p>', unsafe_allow_html=True)
        for uf in uploaded_files:
            file_data[uf.name] = read_file_content(uf)
        # 8대 표준 자료 충족 현황 (파일명 기반 자동 인식)
        _cov_s = coverage_from_filenames([uf.name for uf in uploaded_files])
        st.markdown(material_checklist_html(_cov_s), unsafe_allow_html=True)
    else:
        st.markdown(material_checklist_html(set()), unsafe_allow_html=True)
    if candidate_dept:        file_data["소속 부서"]            = candidate_dept
    if company_standard_s:    file_data["회사 인재상"]           = company_standard_s
    if core_culture_s:        file_data["회사 5대 핵심문화 축"]   = core_culture_s
    if company_direction_s:   file_data["회사 향후 방향성"]       = company_direction_s

    # ── 분석 결과 안내 ──
    st.markdown("""
    <div class="section-header">
        <span class="section-num">03</span>
        <span class="section-title">분석 결과 안내</span>
        <div class="section-rule"></div>
    </div>
    """, unsafe_allow_html=True)
    guide_items = [
        ("#B8924A","01 · 대상자 프로필","종합 한줄 평가와 함께 핵심 성향을 태그로 요약합니다."),
        ("#2B3D5C","02 · 4개 역량 차원","인지능력·잡 전문성·적극성·리더십을 100점 만점으로 점수화합니다."),
        ("#B8924A","03 · 채용 키워드 TOP 3","STAR 기반 면접 질문과 평가 포인트를 순위별로 제공합니다."),
        ("#8B2635","04 · Derailer 위험 요인","스트레스 상황에서 나타날 수 있는 부정적 행동 패턴을 식별합니다."),
        ("#2D6A4F","05 · 성과 극대화 조건","최고 성과를 위한 환경·관리 방식·개발 과제를 제시합니다."),
        ("#2B3D5C","06 · 종합 채용 권고","McKinsey·Korn Ferry 수준의 최종 채용 의사결정 권고를 제공합니다."),
    ]
    g1, g2 = st.columns(2)
    for i, (color, title, desc) in enumerate(guide_items):
        with (g1 if i % 2 == 0 else g2):
            st.markdown(f"""
            <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;
                        padding:1rem 1.2rem;border-left:3px solid {color};margin-bottom:0.8rem;">
                <div style="font-size:0.68rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;
                            color:{color};margin-bottom:0.4rem;">{title}</div>
                <div style="font-size:0.8rem;color:#3D3830;line-height:1.65;">{desc}</div>
            </div>
            """, unsafe_allow_html=True)
    st.markdown('<p style="font-size:0.72rem;color:#B0A898;">※ 업로드된 자료가 많을수록 분석 정확도가 높아집니다.</p>', unsafe_allow_html=True)
    st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

    # ── 아카이브 조회 모드 ──
    if st.session_state.get("show_archive"):
        idx = st.session_state.get("archive_view", 0)
        archive_data = load_archive()
        if 0 <= idx < len(archive_data):
            rec = archive_data[idx]
            st.markdown(f"""
            <div style="background:#EDE8E0;border:1px solid #D4CEC4;border-radius:8px;
                        padding:0.8rem 1.2rem;margin-bottom:1.5rem;">
                <span style="font-size:0.8rem;color:#7A7268;">
                    🗂 아카이브 조회 중 &nbsp;·&nbsp;
                    <b style="color:#1A1714;">{rec.get('candidate_name','')}</b> &nbsp;·&nbsp;
                    {rec.get('saved_at','')}
                </span>
            </div>
            """, unsafe_allow_html=True)
            render_result(rec["result"], rec.get("candidate_name",""))
        if st.button("← 새 분석으로 돌아가기", use_container_width=True, key="s_back"):
            st.session_state["show_archive"] = False
            st.rerun()
    else:
        if st.session_state.get("analysis_done"):
            st.markdown("""
            <div style="background:#FBF8F3;border:1.5px solid #B8924A;border-radius:8px;
                        padding:1rem 1.5rem;margin-bottom:1.2rem;">
                <span style="font-size:0.75rem;font-weight:700;letter-spacing:2px;
                             text-transform:uppercase;color:#B8924A;">분석 완료</span>
                <span style="font-size:0.8rem;color:#7A7268;margin-left:0.8rem;">
                    새 대상자를 분석하려면 아래 버튼을 클릭하세요
                </span>
            </div>
            """, unsafe_allow_html=True)
            if st.button("✦  새 분석 시작 — 초기화", use_container_width=True, key="s_reset"):
                st.session_state["analysis_done"]   = False
                st.session_state["analysis_result"] = None
                st.session_state["analysis_name"]   = None
                st.rerun()
            st.markdown('<div style="height:0.5rem;"></div>', unsafe_allow_html=True)
            st.success("✅ 분석 완료 — 왼쪽 아카이브에 자동 저장되었습니다.")
            render_result(st.session_state["analysis_result"], st.session_state["analysis_name"])
        else:
            run = st.button("◈  분석 시작", use_container_width=True, key="s_run")
            if run:
                if not uploaded_files:
                    st.error("⚠️ 최소 1개 이상의 파일을 업로드해야 분석할 수 있습니다.")
                elif not candidate_name:
                    st.error("⚠️ 검사할 대상자를 선택(또는 입력)해주세요.")
                else:
                    with st.spinner("분석 중 — 업로드된 자료를 종합 검토하고 있습니다..."):
                        try:
                            R = analyze_candidate(api_key, file_data, candidate_name, company_standard_s)
                            R["material_coverage"] = sorted(coverage_from_filenames([uf.name for uf in uploaded_files]))
                            record = {
                                "saved_at":       datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "candidate_name": candidate_name or "이름 없음",
                                "dept":           candidate_dept,
                                "result":         R
                            }
                            save_to_archive(record)
                            st.session_state["analysis_done"]   = True
                            st.session_state["analysis_result"] = R
                            st.session_state["analysis_name"]   = candidate_name
                            st.rerun()
                        except json.JSONDecodeError as e:
                            st.error(f"결과 파싱 오류. 잠시 후 다시 시도해주세요. (상세: {str(e)[:80]})")
                        except anthropic.AuthenticationError:
                            st.error("API Key가 유효하지 않습니다.")
                        except anthropic.APIStatusError as e:
                            st.error(f"API 오류 ({e.status_code}): {str(e.message)[:120]}")
                        except Exception as e:
                            st.error(f"오류 발생: {e}")


# ════════════════════════════════════════════════════════
#  TAB 2 — 대량 분석
# ════════════════════════════════════════════════════════
with tab_bulk:
    sub_manual, sub_auto = st.tabs(["✍️  수동 등록", "⚡  자동 등록 (파일명 자동 인식)"])

    with sub_manual:
        st.markdown("""
        <div class="section-header">
            <span class="section-num">01</span>
            <span class="section-title">대량 분석 대상자 등록</span>
            <div class="section-rule"></div>
        </div>
        <p style="font-size:0.8rem;color:#7A7268;margin-bottom:1.2rem;">
            조직 명부에서 <b>여러 명을 한 번에 선택</b>해 목록에 담고, 대상자별 자료를 올린 뒤
            <b>전체 분석 시작</b>을 누르면 순차 분석됩니다. 결과는 아카이브에 자동 저장돼요.
        </p>
        """, unsafe_allow_html=True)

        if "bulk_list"    not in st.session_state: st.session_state["bulk_list"]    = []
        if "bulk_results" not in st.session_state: st.session_state["bulk_results"] = []
        if "bulk_done"    not in st.session_state: st.session_state["bulk_done"]    = False

        _org_b = load_org_data()
        _done_b = {r.get("candidate_name", "") for r in load_archive() if r.get("candidate_name")}

        def _flatten_b(org):
            out = {}
            for bonbu, bd in org.items():
                units = {}
                bdir = bd.get("_직속", {})
                dp = bdir.get("_직속", []) if isinstance(bdir, dict) else []
                if dp:
                    units["(본부 직속)"] = dp
                for tname, tdict in bd.items():
                    if tname == "_직속":
                        continue
                    td = tdict.get("_직속", [])
                    if td:
                        units[tname] = td
                    for pname, plist in tdict.items():
                        if pname == "_직속":
                            continue
                        units[f"{tname} · {pname}"] = plist
                if units:
                    out[bonbu] = units
            return out

        # ── 대상자 추가 (조직 명부에서 체크박스로 여러 명 선택) ──
        with st.expander("➕  조직 명부에서 대상자 추가 (체크박스로 여러 명 선택)", expanded=not st.session_state["bulk_done"]):
            _existing = {e["name"] for e in st.session_state["bulk_list"]}
            if _org_b:
                _flatb = _flatten_b(_org_b)
                bc1, bc2 = st.columns([1, 1.6])
                with bc1:
                    bb = st.selectbox("본부", ["전체"] + list(_flatb.keys()), key="b_bonbu")

                # 팀·파트 선택지 구성 (팀 선택 시 하위 파트 인원까지 포함)
                _optmembers = {}   # 라벨 -> [(person, dept)]
                _opts = []
                if bb != "전체":
                    _bunits = _flatb.get(bb, {})
                    _opts.append("(본부 전체)")
                    _optmembers["(본부 전체)"] = [(p, f"{bb} / {lbl}") for lbl, lst in _bunits.items() for p in lst]
                    if "(본부 직속)" in _bunits:
                        _opts.append("(본부 직속)")
                        _optmembers["(본부 직속)"] = [(p, f"{bb} / (본부 직속)") for p in _bunits["(본부 직속)"]]
                    _teams, _seen = [], set()
                    for lbl in _bunits:
                        if lbl == "(본부 직속)":
                            continue
                        t = lbl.split(" · ", 1)[0] if " · " in lbl else lbl
                        if t not in _seen:
                            _seen.add(t)
                            _teams.append(t)
                    for t in _teams:
                        _direct = _bunits.get(t, [])
                        _plabels = [lbl for lbl in _bunits if lbl.startswith(t + " · ")]
                        _team_all = [(p, f"{bb} / {t}") for p in _direct]
                        for pl in _plabels:
                            _team_all += [(p, f"{bb} / {pl}") for p in _bunits[pl]]
                        if _plabels:
                            _tl = f"{t} (팀 전체 · 하위 파트 포함)"
                            _opts.append(_tl)
                            _optmembers[_tl] = _team_all
                            if _direct:
                                _dl = f"{t} · (팀 직속)"
                                _opts.append(_dl)
                                _optmembers[_dl] = [(p, f"{bb} / {t}") for p in _direct]
                            for pl in _plabels:
                                _opts.append(pl)
                                _optmembers[pl] = [(p, f"{bb} / {pl}") for p in _bunits[pl]]
                        else:
                            _opts.append(t)
                            _optmembers[t] = _team_all

                with bc2:
                    if bb != "전체":
                        bu = st.selectbox("팀 · 파트", _opts, key="b_unit")
                    else:
                        bu = None
                        st.markdown('<div style="font-size:0.8rem;color:#7A7268;padding-top:1.9rem;">전체 본부의 인원을 한 화면에서 선택</div>', unsafe_allow_html=True)
                bsearch = st.text_input("이름 검색 (선택)", key="b_search", placeholder="이름 일부를 입력하면 목록이 좁혀집니다")

                # 후보 인원 수집 (사람, 소속 라벨)
                _cands = []
                if bb == "전체":
                    for _bn, _us in _flatb.items():
                        for _u, _lst in _us.items():
                            for p in _lst:
                                _cands.append((p, f"{_bn} / {_u}"))
                else:
                    _cands = list(_optmembers.get(bu, []))
                if bsearch and bsearch.strip():
                    _cands = [(p, d) for (p, d) in _cands if bsearch.strip() in p["name"]]

                _selectable = [(p, d) for (p, d) in _cands if p["name"] not in _existing]
                st.markdown(
                    f'<div style="font-size:0.8rem;color:#7A7268;margin:0.5rem 0 0.4rem;">'
                    f'표시 <b style="color:#1A1714;">{len(_cands)}</b>명 · 추가 가능 <b style="color:#1A1714;">{len(_selectable)}</b>명 '
                    f'<span style="color:#B0A898;">(✅ 분석 등록 / ⚪ 미등록 · 이미 담은 사람은 비활성)</span></div>',
                    unsafe_allow_html=True)

                sc1, sc2 = st.columns(2)
                with sc1:
                    if st.button("표시된 인원 전체 선택", use_container_width=True, key="b_selall"):
                        for p, d in _selectable:
                            st.session_state[f"bchk_{p['name']}"] = True
                        st.rerun()
                with sc2:
                    if st.button("선택 모두 해제", use_container_width=True, key="b_clrall"):
                        for p, d in _cands:
                            st.session_state[f"bchk_{p['name']}"] = False
                        st.rerun()

                if _cands:
                    _ck_cols = st.columns(3)
                    for _idx, (p, d) in enumerate(_cands):
                        _tag = "✅" if p["name"] in _done_b else "⚪"
                        with _ck_cols[_idx % 3]:
                            if p["name"] in _existing:
                                st.checkbox(f"{_tag} {p['name']} · 담음", value=True, disabled=True, key=f"bchkD_{p['name']}")
                            else:
                                st.checkbox(f"{_tag} {p['name']} · {p.get('pos','')}", key=f"bchk_{p['name']}")
                else:
                    st.caption("조건에 맞는 인원이 없습니다.")

                if st.button("✓ 선택한 인원 목록에 추가", use_container_width=True, key="b_add"):
                    added = 0
                    cur = {e["name"] for e in st.session_state["bulk_list"]}
                    _dmap = {p["name"]: d for p, d in _cands}
                    for p, d in _selectable:
                        if p["name"] in cur:
                            continue
                        if st.session_state.get(f"bchk_{p['name']}"):
                            st.session_state["bulk_list"].append({
                                "name": p["name"], "dept": _dmap.get(p["name"], ""),
                                "file_data": {}, "file_count": 0})
                            cur.add(p["name"])
                            added += 1
                    if added:
                        st.success(f"✅ {added}명 목록에 추가됨")
                        st.rerun()
                    else:
                        st.warning("체크된 새 인원이 없습니다.")
            else:
                st.info("org_data.json이 없어 명부 선택을 사용할 수 없어요. 아래 직접 입력을 이용하세요.")

            st.markdown('<div style="height:0.3rem;"></div>', unsafe_allow_html=True)
            with st.expander("✏️ 직접 입력으로 추가 (명부에 없는 경우)"):
                mc1, mc2 = st.columns(2)
                with mc1: man_name = st.text_input("성명", key="b_manname")
                with mc2: man_dept = st.text_input("소속", key="b_mandept")
                if st.button("이 인원 추가", key="b_manadd"):
                    if not man_name:
                        st.error("성명을 입력해주세요.")
                    elif man_name in {e["name"] for e in st.session_state["bulk_list"]}:
                        st.warning("이미 목록에 있습니다.")
                    else:
                        st.session_state["bulk_list"].append({"name": man_name, "dept": man_dept, "file_data": {}, "file_count": 0})
                        st.success(f"✅ {man_name} 추가됨")
                        st.rerun()

        # ── 등록된 목록 · 대상자별 자료 업로드 ──
        if st.session_state["bulk_list"]:
            st.markdown("""
            <div class="section-header" style="margin-top:1.5rem;">
                <span class="section-num">02</span>
                <span class="section-title">등록된 대상자 · 자료 업로드</span>
                <div class="section-rule"></div>
            </div>
            <p style="font-size:0.76rem;color:#7A7268;margin-bottom:0.8rem;">
                각 대상자에 8대 표준 자료(이력서·포트폴리오·다면평가·기안서·MBTI·인적성·SNS·기타)를 올려주세요.
                파일명에 자료 종류를 넣으면 충족 항목이 자동 인식됩니다. 자료가 없는 대상자는 분석에서 제외됩니다.
            </p>
            """, unsafe_allow_html=True)
            st.markdown(material_checklist_html(set(), title="8대 표준 분석자료 항목"), unsafe_allow_html=True)

            pending_files = {}
            for i, cand in enumerate(st.session_state["bulk_list"]):
                done = i < len(st.session_state["bulk_results"]) and st.session_state["bulk_results"][i].get("success")
                reg = "✅" if cand["name"] in _done_b else "⚪"
                lc1, lc2 = st.columns([5, 1])
                with lc1:
                    stat = "✅ 분석 완료" if done else "⏳ 대기"
                    st.markdown(f"""
                    <div style="background:white;border:1px solid #D4CEC4;border-radius:6px;
                                padding:0.6rem 1rem;border-left:3px solid {'#2D6A4F' if done else '#B8924A'};">
                        <span style="font-size:0.95rem;">{reg}</span>
                        <b style="font-size:0.9rem;color:#1A1714;margin-left:0.3rem;">{cand['name']}</b>
                        <span style="font-size:0.74rem;color:#7A7268;margin-left:0.6rem;">{cand.get('dept','')}</span>
                        <span style="font-size:0.72rem;color:#B0A898;margin-left:0.6rem;">{stat}</span>
                    </div>
                    """, unsafe_allow_html=True)
                with lc2:
                    if not st.session_state["bulk_done"]:
                        if st.button("삭제", key=f"bdel_{i}", use_container_width=True):
                            st.session_state["bulk_list"].pop(i)
                            if i < len(st.session_state["bulk_results"]):
                                st.session_state["bulk_results"].pop(i)
                            st.rerun()
                if not st.session_state["bulk_done"]:
                    up = st.file_uploader(
                        f"{cand['name']} 자료",
                        type=["pdf", "docx", "jpg", "jpeg", "png", "webp", "txt", "md"],
                        accept_multiple_files=True, key=f"bup_{cand['name']}",
                        label_visibility="collapsed"
                    )
                    pending_files[i] = up
                    st.caption(f"📎 {cand['name']} — {len(up) if up else 0}개 업로드됨")
                    _cov_b = coverage_from_filenames([u.name for u in up]) if up else set()
                    st.markdown(material_chips_inline(_cov_b), unsafe_allow_html=True)

            st.markdown('<div style="height:0.8rem;"></div>', unsafe_allow_html=True)

            # ── 전체 분석 시작 ──
            if not st.session_state["bulk_done"]:
                _ready = sum(1 for i in pending_files if pending_files[i])
                if st.button(f"◈  전체 분석 시작 (자료 등록 {_ready}명)", use_container_width=True, key="b_run"):
                    if _ready == 0:
                        st.error("⚠️ 자료가 업로드된 대상자가 없습니다. 각 대상자에 자료를 올려주세요.")
                    else:
                        st.session_state["bulk_results"] = []
                        n_t = len(st.session_state["bulk_list"])
                        prog = st.progress(0, text="분석 준비 중...")
                        for k, cand in enumerate(st.session_state["bulk_list"]):
                            prog.progress(k / n_t, text=f"({k+1}/{n_t}) {cand['name']} ...")
                            ups = pending_files.get(k)
                            if not ups:
                                st.session_state["bulk_results"].append({
                                    "name": cand["name"], "dept": cand.get("dept", ""),
                                    "result": {}, "success": False, "error": "자료 미업로드"})
                                prog.progress((k + 1) / n_t)
                                continue
                            try:
                                fd = {uf.name: read_file_content(uf) for uf in ups}
                                if cand.get("dept"): fd["소속 부서"] = cand["dept"]
                                fd["회사 인재상"] = company_standard
                                fd["회사 5대 핵심문화 축"] = core_culture
                                fd["회사 향후 방향성"] = company_direction
                                R = analyze_candidate(api_key, fd, cand["name"], company_standard)
                                R["material_coverage"] = sorted(coverage_from_filenames([uf.name for uf in ups]))
                                st.session_state["bulk_results"].append({
                                    "name": cand["name"], "dept": cand.get("dept", ""),
                                    "result": R, "success": True})
                                save_to_archive({
                                    "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                    "candidate_name": cand["name"], "dept": cand.get("dept", ""),
                                    "result": R})
                            except Exception as e:
                                st.session_state["bulk_results"].append({
                                    "name": cand["name"], "dept": cand.get("dept", ""),
                                    "result": {}, "success": False, "error": str(e)})
                            prog.progress((k + 1) / n_t, text=f"({k+1}/{n_t}) {cand['name']} 완료")
                        prog.progress(1.0, text="✅ 분석 완료!")
                        st.session_state["bulk_done"] = True
                        st.rerun()

            # ── 대량 분석 결과 ──
            if st.session_state["bulk_done"] and st.session_state["bulk_results"]:
                st.markdown("""
                <div class="section-header" style="margin-top:2rem;">
                    <span class="section-num">03</span>
                    <span class="section-title">대량 분석 결과 요약</span>
                    <div class="section-rule"></div>
                </div>
                """, unsafe_allow_html=True)

                # 비교 테이블 (종합 점수 순위 + 리밸런싱 판정 포함)
                dim_keys  = ["cognitive_ability","job_expertise","proactiveness","leadership"]
                dim_names = ["인지","전문성","적극성","리더십"]
                re_emoji  = {"LOW":"🟢","MEDIUM":"🟡","HIGH":"🔴","CRITICAL":"🔴"}.get
                verdict_emoji = {"KEEP":"✅","DEVELOP":"📈","WATCH":"👁","MISFIT":"⚠️"}.get

                # 종합 점수 계산 후 내림차순 정렬 (분석 실패 건은 맨 아래)
                scored = []
                for br in st.session_state["bulk_results"]:
                    ov = compute_overall_score(br["result"])[0] if br["success"] else None
                    scored.append((ov if ov is not None else -1, br, ov))
                scored.sort(key=lambda x: x[0], reverse=True)

                rows_header = "| 순위 | 종합 | 이름 | 부서 | 판정 | 조직적합 | 리더십 | " + " | ".join(dim_names) + " | 번아웃 | 이직 |"
                _ncols = rows_header.count("|") - 1
                rows = [rows_header, "|" + "|".join([":---:"] * _ncols) + "|"]
                rank = 0
                for _, br, ov in scored:
                    if not br["success"]:
                        rows.append(f"| - | - | {br['name']} | {br['dept']} | ❌ | | | | | | | | |")
                        continue
                    rank += 1
                    medal = {1:"🥇",2:"🥈",3:"🥉"}.get(rank, str(rank))
                    R = br["result"]
                    dims    = R.get("dimensions", {})
                    scores  = [str(dims.get(k,{}).get("score","?")) for k in dim_keys]
                    b_lvl   = R.get("burnout_risk",{}).get("level","?")
                    t_lvl   = R.get("turnover_risk",{}).get("level","?")
                    verdict = R.get("rebalancing_verdict",{}).get("decision","?")
                    of_sc   = R.get("org_fit",{}).get("score","?")
                    lr_sc   = R.get("leadership_readiness",{}).get("score","?")
                    rows.append(f"| {medal} | **{ov}** | **{br['name']}** | {br['dept']} | {verdict_emoji(verdict,'⚪')} {verdict} | {of_sc} | {lr_sc} | {' | '.join(scores)} | {re_emoji(b_lvl,'⚪')} | {re_emoji(t_lvl,'⚪')} |")
                st.markdown("\n".join(rows))

                st.markdown("""
                <p style="font-size:0.7rem;color:#B0A898;margin-top:0.5rem;">
                종합 점수 높은 순으로 정렬 · 종합 = 세부 점수 가중합(역량35%·조직적합30%·리더십준비15%·저위험20%)
                <br>판정: ✅ KEEP(핵심·유지) · 📈 DEVELOP(육성) · 👁 WATCH(관찰) · ⚠️ MISFIT(부적합)
                &nbsp;|&nbsp; 모든 점수 100점 만점
                </p>
                """, unsafe_allow_html=True)

                # ── 집단 비교 분석 ──
                st.markdown("""
                <div class="section-header" style="margin-top:2rem;">
                    <span class="section-num">04</span>
                    <span class="section-title">집단 비교 분석</span>
                    <div class="section-rule"></div>
                </div>
                <p style="font-size:0.78rem;color:#7A7268;margin-bottom:1rem;">
                    부서·판정 분포를 한눈에 파악합니다.
                </p>
                """, unsafe_allow_html=True)

                ok_results = [br for br in st.session_state["bulk_results"] if br["success"]]

                # 판정 분포
                verdict_count = {}
                for br in ok_results:
                    v = br["result"].get("rebalancing_verdict",{}).get("decision","미분류")
                    verdict_count[v] = verdict_count.get(v, 0) + 1

                vc1, vc2, vc3, vc4 = st.columns(4)
                v_meta = [("KEEP","✅","#2D6A4F"),("DEVELOP","📈","#2B3D5C"),("WATCH","👁","#8B6914"),("MISFIT","⚠️","#8B2635")]
                for col, (vk, emoji, vcolor) in zip([vc1,vc2,vc3,vc4], v_meta):
                    with col:
                        cnt = verdict_count.get(vk, 0)
                        st.markdown(f"""
                        <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;
                                    padding:1.2rem;text-align:center;border-top:3px solid {vcolor};">
                            <div style="font-size:1.5rem;">{emoji}</div>
                            <div style="font-family:'DM Serif Display',serif;font-size:2rem;
                                        color:{vcolor};font-style:italic;">{cnt}</div>
                            <div style="font-size:0.65rem;letter-spacing:1px;color:#7A7268;
                                        text-transform:uppercase;">{vk}</div>
                        </div>
                        """, unsafe_allow_html=True)

                # 부서별 평균
                from collections import defaultdict
                dept_data = defaultdict(list)
                for br in ok_results:
                    dept_data[br["dept"] or "미지정"].append(br["result"])

                if len(dept_data) > 1:
                    st.markdown('<div style="height:1rem;"></div>', unsafe_allow_html=True)
                    st.markdown('<p style="font-size:0.75rem;font-weight:600;color:#3D3830;margin-bottom:0.5rem;">📊 부서별 평균 역량</p>', unsafe_allow_html=True)
                    drows = ["| 부서 | 인원 | 종합 | 인지 | 전문성 | 적극성 | 리더십 | 조직적합 |",
                             "|------|:---:|:---:|:---:|:---:|:---:|:---:|:---:|"]
                    for dept, results in dept_data.items():
                        n = len(results)
                        def avg(key, sub="score"):
                            vals = [r.get("dimensions",{}).get(key,{}).get(sub,0) for r in results]
                            vals = [v for v in vals if isinstance(v,(int,float))]
                            return round(sum(vals)/len(vals)) if vals else "-"
                        of_vals = [r.get("org_fit",{}).get("score",0) for r in results]
                        of_vals = [v for v in of_vals if isinstance(v,(int,float))]
                        of_avg = round(sum(of_vals)/len(of_vals)) if of_vals else "-"
                        ov_vals = [compute_overall_score(r)[0] for r in results]
                        ov_vals = [v for v in ov_vals if isinstance(v,(int,float))]
                        ov_avg = round(sum(ov_vals)/len(ov_vals)) if ov_vals else "-"
                        drows.append(f"| {dept} | {n} | **{ov_avg}** | {avg('cognitive_ability')} | {avg('job_expertise')} | {avg('proactiveness')} | {avg('leadership')} | {of_avg} |")
                    st.markdown("\n".join(drows))

                # ── 리더 적합성 스크리닝 ──
                st.markdown("""
                <div class="section-header" style="margin-top:1.8rem;">
                    <span class="section-num">05</span>
                    <span class="section-title">리더 적합성 스크리닝</span>
                    <div class="section-rule"></div>
                </div>
                """, unsafe_allow_html=True)
                lead_buckets = {"즉시 가능 (80+)": [], "육성 후 가능 (60–79)": [], "현재 부적합 (<60)": []}
                for br in ok_results:
                    lr = br["result"].get("leadership_readiness", {}).get("score")
                    if not isinstance(lr, (int, float)):
                        continue
                    if lr >= 80:   lead_buckets["즉시 가능 (80+)"].append((br["name"], lr))
                    elif lr >= 60: lead_buckets["육성 후 가능 (60–79)"].append((br["name"], lr))
                    else:          lead_buckets["현재 부적합 (<60)"].append((br["name"], lr))
                lb_cols = st.columns(3)
                lb_meta = [("즉시 가능 (80+)", "#2D6A4F"), ("육성 후 가능 (60–79)", "#2B3D5C"), ("현재 부적합 (<60)", "#8B2635")]
                for col, (label, color) in zip(lb_cols, lb_meta):
                    people = sorted(lead_buckets[label], key=lambda x: x[1], reverse=True)
                    names_html = "".join(
                        f'<div style="font-size:0.75rem;color:#3D3830;margin:2px 0;">{n} <b style="color:{color};">{s}</b></div>'
                        for n, s in people
                    ) or '<div style="font-size:0.72rem;color:#B0A898;">해당 없음</div>'
                    with col:
                        st.markdown(f"""
                        <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;padding:1rem 1.1rem;border-top:3px solid {color};">
                            <div style="font-size:0.64rem;font-weight:700;letter-spacing:1px;color:{color};text-transform:uppercase;margin-bottom:0.5rem;">{label} · {len(people)}명</div>
                            {names_html}
                        </div>
                        """, unsafe_allow_html=True)

                # ── MISFIT·WATCH 공통 특징 추출 ──
                from collections import Counter
                risk_people = [br for br in ok_results
                               if br["result"].get("rebalancing_verdict", {}).get("decision") in ("MISFIT", "WATCH")]
                if risk_people:
                    tag_counter = Counter()
                    for br in risk_people:
                        for t in br["result"].get("personality_tags", []):
                            tag_counter[t] += 1
                    common = [(t, c) for t, c in tag_counter.most_common(8) if c >= 2]
                    def _avg_field(people, getter):
                        vals = [getter(br["result"]) for br in people]
                        vals = [v for v in vals if isinstance(v, (int, float))]
                        return round(sum(vals) / len(vals)) if vals else "-"
                    avg_of = _avg_field(risk_people, lambda R: R.get("org_fit", {}).get("score"))
                    avg_ov = _avg_field(risk_people, lambda R: compute_overall_score(R)[0])
                    names = ", ".join(br["name"] for br in risk_people)
                    tags_html = "".join(
                        f'<span style="display:inline-block;background:#FAEAEC;border:1px solid #E6C9CE;border-radius:4px;padding:2px 9px;margin:2px;font-size:0.72rem;color:#8B2635;">{t} ×{c}</span>'
                        for t, c in common
                    ) or '<span style="font-size:0.73rem;color:#B0A898;">2명 이상이 공유하는 공통 태그 없음</span>'
                    st.markdown("""
                    <div class="section-header" style="margin-top:1.8rem;">
                        <span class="section-num">06</span>
                        <span class="section-title">MISFIT · WATCH 공통 특징 (제외 기준 패턴)</span>
                        <div class="section-rule"></div>
                    </div>
                    """, unsafe_allow_html=True)
                    st.markdown(f"""
                    <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;padding:1.2rem 1.4rem;border-left:3px solid #8B2635;">
                        <div style="font-size:0.74rem;color:#7A7268;margin-bottom:0.6rem;">대상 {len(risk_people)}명 · 평균 종합 {avg_ov} · 평균 조직적합 {avg_of}</div>
                        <div style="font-size:0.7rem;color:#B0A898;margin-bottom:0.4rem;">공유 성향 태그</div>
                        <div>{tags_html}</div>
                        <div style="font-size:0.7rem;color:#7A7268;margin-top:0.7rem;">대상자: {names}</div>
                    </div>
                    <p style="font-size:0.68rem;color:#B0A898;margin-top:0.5rem;">※ 자동 추출된 통계 패턴입니다. 개인 인사 판단은 반드시 추가 검토를 거치세요.</p>
                    """, unsafe_allow_html=True)

                # ── 집단 비교 (전공·대학·학력·출신지역) ──
                def group_overall_table(field_label, getter):
                    groups = defaultdict(list)
                    for br in ok_results:
                        key = getter(br["result"]) or "미상"
                        if not key or key == "자료 미제공":
                            key = "미상"
                        ov = compute_overall_score(br["result"])[0]
                        if isinstance(ov, (int, float)):
                            groups[key].append(ov)
                    if not [k for k in groups if k != "미상"]:
                        return None
                    rows = [f"| {field_label} | 인원 | 평균 종합 |", "|------|:---:|:---:|"]
                    for k, vals in sorted(groups.items(), key=lambda kv: sum(kv[1]) / len(kv[1]), reverse=True):
                        rows.append(f"| {k} | {len(vals)} | **{round(sum(vals)/len(vals))}** |")
                    return "\n".join(rows)

                group_specs = [
                    ("학력수준", lambda R: R.get("profile", {}).get("education_level")),
                    ("전공",     lambda R: R.get("profile", {}).get("major")),
                    ("대학",     lambda R: R.get("profile", {}).get("university")),
                    ("출신지역", lambda R: R.get("profile", {}).get("region")),
                ]
                group_tables = [(lbl, group_overall_table(lbl, g)) for lbl, g in group_specs]
                group_tables = [(lbl, t) for lbl, t in group_tables if t]
                if group_tables:
                    st.markdown("""
                    <div class="section-header" style="margin-top:1.8rem;">
                        <span class="section-num">07</span>
                        <span class="section-title">집단 비교 · 전공·대학·학력·지역</span>
                        <div class="section-rule"></div>
                    </div>
                    <p style="font-size:0.72rem;color:#B0A898;margin-bottom:0.6rem;">
                        프로필이 추출된 인원만 집계됩니다(미추출은 '미상'). '학력수준'으로 고학력=고성과 여부를 가늠할 수 있습니다.
                        출신지역은 통계 참고용이며 개인 점수에는 반영되지 않습니다.
                    </p>
                    """, unsafe_allow_html=True)
                    for lbl, t in group_tables:
                        st.markdown(f'<p style="font-size:0.74rem;font-weight:600;color:#3D3830;margin:0.7rem 0 0.3rem;">📊 {lbl}별 평균 종합 점수</p>', unsafe_allow_html=True)
                        st.markdown(t)

                st.markdown('<div style="height:1.5rem;"></div>', unsafe_allow_html=True)

                # ── 다운로드 (JSON + Excel) ──
                dl1, dl2 = st.columns(2)
                with dl1:
                    export_bulk = [{"name":br["name"],"dept":br["dept"],"result":br.get("result",{})} for br in ok_results]
                    st.download_button(
                        "⬇ 전체 결과 JSON",
                        data=json.dumps(export_bulk, ensure_ascii=False, indent=2).encode("utf-8"),
                        file_name=f"bulk_analysis_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
                        mime="application/json", use_container_width=True
                    )
                with dl2:
                    # Excel 생성
                    try:
                        import io as _io
                        import openpyxl
                        from openpyxl.styles import Font, PatternFill, Alignment

                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.title = "인재분석결과"
                        headers = ["순위","이름","부서","종합점수","리밸런싱판정","신뢰도","조직적합도","방향성적합도","리더십준비도",
                                   "인지능력","잡전문성","적극성","리더십",
                                   "번아웃","이직위험","SNS점수","전공","대학","학력수준","출신지역",
                                   "학력성과정합","추천커리어트랙","한줄평","리밸런싱근거"]
                        ws.append(headers)
                        for c in range(1, len(headers)+1):
                            cell = ws.cell(row=1, column=c)
                            cell.font = Font(bold=True, color="FFFFFF", size=10)
                            cell.fill = PatternFill("solid", fgColor="1A1714")
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                        # 종합 점수 내림차순 정렬 후 순위 부여
                        excel_sorted = sorted(
                            ok_results,
                            key=lambda br: (compute_overall_score(br["result"])[0] or 0),
                            reverse=True
                        )
                        for rank_i, br in enumerate(excel_sorted, 1):
                            R = br["result"]
                            d = R.get("dimensions",{})
                            p = R.get("profile",{})
                            sn = R.get("sns_analysis",{})
                            ov = compute_overall_score(R)[0]
                            ws.append([
                                rank_i,
                                br["name"], br["dept"],
                                ov if ov is not None else "",
                                R.get("rebalancing_verdict",{}).get("decision",""),
                                R.get("rebalancing_verdict",{}).get("confidence",""),
                                R.get("org_fit",{}).get("score",""),
                                R.get("direction_fit",{}).get("score",""),
                                R.get("leadership_readiness",{}).get("score",""),
                                d.get("cognitive_ability",{}).get("score",""),
                                d.get("job_expertise",{}).get("score",""),
                                d.get("proactiveness",{}).get("score",""),
                                d.get("leadership",{}).get("score",""),
                                R.get("burnout_risk",{}).get("level",""),
                                R.get("turnover_risk",{}).get("level",""),
                                sn.get("score","") if sn.get("available") else "",
                                p.get("major",""),
                                p.get("university",""),
                                p.get("education_level",""),
                                p.get("region",""),
                                R.get("credential_performance",{}).get("alignment",""),
                                R.get("career_track",{}).get("recommended_track",""),
                                R.get("candidate_summary",""),
                                R.get("rebalancing_verdict",{}).get("rationale",""),
                            ])
                        # 열 너비
                        widths = [6,10,16,8,12,7,8,9,9,8,8,8,8,8,8,8,12,12,9,10,11,14,36,46]
                        for i, w in enumerate(widths, 1):
                            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

                        buf = _io.BytesIO()
                        wb.save(buf)
                        st.download_button(
                            "⬇ 엑셀(.xlsx) 다운로드",
                            data=buf.getvalue(),
                            file_name=f"인재분석_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    except Exception as e:
                        st.caption(f"엑셀 생성 오류: {e}")

                st.markdown('<div style="height:1.5rem;"></div>', unsafe_allow_html=True)

                # 개별 상세 결과
                st.markdown("""
                <div class="section-header">
                    <span class="section-num">08</span>
                    <span class="section-title">개별 상세 결과</span>
                    <div class="section-rule"></div>
                </div>
                """, unsafe_allow_html=True)
                for br in st.session_state["bulk_results"]:
                    icon = "✅" if br["success"] else "❌"
                    with st.expander(f"{icon}  {br['name']} · {br.get('dept','')}"):
                        if br["success"]:
                            render_result(br["result"], br["name"])
                        else:
                            st.error(f"분석 실패: {br.get('error','')}")

                if st.button("🔄  초기화 (새 대량 분석 시작)", use_container_width=True, key="b_reset"):
                    st.session_state["bulk_list"]    = []
                    st.session_state["bulk_results"] = []
                    st.session_state["bulk_done"]    = False
                    st.rerun()
        else:
            st.markdown('<p style="font-size:0.82rem;color:#B0A898;text-align:center;padding:2rem 0;">위 폼에서 대상자를 추가해주세요.</p>', unsafe_allow_html=True)

    with sub_auto:
        st.markdown("""
        <div class="section-header">
            <span class="section-num">01</span>
            <span class="section-title">자료 일괄 업로드 · 파일명 자동 인식</span>
            <div class="section-rule"></div>
        </div>
        <p style="font-size:0.8rem;color:#7A7268;margin-bottom:1rem;">
            여러 자료를 한꺼번에 올리면 <b>파일명에 들어있는 이름</b>을 인식해 사람별로 자동 분류·등록합니다.
            예: <code>홍길동_이력서.pdf</code>, <code>홍길동 MBTI.png</code>, <code>김철수-자기소개서.docx</code>.
            등록 내용을 확인한 뒤 <b>분석 시작</b>만 누르면 일괄 분석돼요.
        </p>
        """, unsafe_allow_html=True)

        if "auto_results" not in st.session_state: st.session_state["auto_results"] = []
        if "auto_done"    not in st.session_state: st.session_state["auto_done"]    = False
        if "auto_excluded" not in st.session_state: st.session_state["auto_excluded"] = set()

        _orgA = load_org_data()
        _doneA = {r.get("candidate_name", "") for r in load_archive() if r.get("candidate_name")}
        _deptA = {}
        _rosterA = []
        if _orgA:
            for _bn, _bd in _orgA.items():
                _bdir = _bd.get("_직속", {})
                for _p in (_bdir.get("_직속", []) if isinstance(_bdir, dict) else []):
                    _deptA[_p["name"]] = f"{_bn} / (본부 직속)"
                for _tn, _td in _bd.items():
                    if _tn == "_직속":
                        continue
                    for _p in _td.get("_직속", []):
                        _deptA[_p["name"]] = f"{_bn} / {_tn}"
                    for _pn, _pl in _td.items():
                        if _pn == "_직속":
                            continue
                        for _p in _pl:
                            _deptA[_p["name"]] = f"{_bn} / {_tn} · {_pn}"
            _rosterA = sorted(_deptA.keys(), key=len, reverse=True)  # 긴 이름 우선 매칭

        import re as _re
        def _match_name(fn):
            base = fn.rsplit(".", 1)[0]
            for nm in _rosterA:
                if nm and nm in base:
                    return nm
            for tok in _re.split(r"[\s_\-.()\[\]]+", base):
                if _re.fullmatch(r"[가-힣]{2,4}", tok):
                    return tok
            return None

        auto_files = st.file_uploader(
            "자료 일괄 업로드 (여러 파일 동시 선택)",
            type=["pdf", "docx", "jpg", "jpeg", "png", "webp", "txt", "md"],
            accept_multiple_files=True, key="auto_uploader")

        if auto_files and not st.session_state["auto_done"]:
            grouped, unmatched, excluded = {}, [], {}
            for f in auto_files:
                nm = _match_name(f.name)
                if not nm:
                    unmatched.append(f)
                elif nm in st.session_state["auto_excluded"]:
                    excluded.setdefault(nm, []).append(f)
                else:
                    grouped.setdefault(nm, []).append(f)
            names_sorted = sorted(grouped.keys())

            st.markdown(
                f'<div style="display:flex;gap:1.2rem;flex-wrap:wrap;margin:0.6rem 0 0.8rem;font-size:0.82rem;color:#7A7268;">'
                f'<span>업로드 파일 <b style="color:#1A1714;">{len(auto_files)}</b>개</span>'
                f'<span>인식된 대상자 <b style="color:#2D6A4F;">{len(names_sorted)}</b>명</span>'
                f'<span>제외됨 <b style="color:#8B6914;">{len(excluded)}</b>건</span>'
                f'<span>인식 실패 <b style="color:#B0392B;">{len(unmatched)}</b>개</span></div>',
                unsafe_allow_html=True)

            st.markdown("""
            <div class="section-header" style="margin-top:0.6rem;">
                <span class="section-num">02</span>
                <span class="section-title">자동 등록된 대상자 · 자료 확인</span>
                <div class="section-rule"></div>
            </div>
            <p style="font-size:0.74rem;color:#7A7268;margin-bottom:0.6rem;">
                사람이 아닌 항목(예: ‘보고서’, 회사명 등)이 잘못 인식된 경우 오른쪽 <b>제외</b> 버튼으로 분석 대상에서 빼주세요.
            </p>
            """, unsafe_allow_html=True)

            for nm in names_sorted:
                fs = grouped[nm]
                tag = "✅" if nm in _doneA else "⚪"
                dept = _deptA.get(nm, "(명부 외)")
                fnames = " · ".join(f.name for f in fs)
                _cov_a = coverage_from_filenames([f.name for f in fs])
                _suspect = dept == "(명부 외)"
                cc1, cc2 = st.columns([6, 1])
                with cc1:
                    st.markdown(
                        f'<div style="background:white;border:1px solid {"#E4C98A" if _suspect else "#D4CEC4"};border-radius:6px;'
                        f'padding:0.6rem 1rem;border-left:3px solid {"#C9A227" if _suspect else "#B8924A"};margin-bottom:0.1rem;">'
                        f'<span style="font-size:0.95rem;">{tag}</span> '
                        f'<b style="font-size:0.92rem;color:#1A1714;">{nm}</b> '
                        f'<span style="font-size:0.74rem;color:#7A7268;">· {dept}</span> '
                        f'<span style="font-size:0.74rem;color:#2D6A4F;font-weight:700;">· {len(fs)}개 자료</span>'
                        f'{" <span style=\'font-size:0.7rem;color:#8B6914;\'>· ⚠️ 명부 외 — 확인 필요</span>" if _suspect else ""}'
                        f'<div style="font-size:0.72rem;color:#B0A898;margin-top:0.25rem;">📎 {fnames}</div>'
                        f'{material_chips_inline(_cov_a)}</div>',
                        unsafe_allow_html=True)
                with cc2:
                    if st.button("✕ 제외", key=f"auto_del_{nm}", use_container_width=True):
                        st.session_state["auto_excluded"].add(nm)
                        st.rerun()

            if excluded:
                with st.expander(f"🚫 제외된 항목 {len(excluded)}건 (분석 대상에서 빠짐 · 복원 가능)", expanded=False):
                    for nm in sorted(excluded.keys()):
                        ec1, ec2 = st.columns([6, 1])
                        with ec1:
                            st.markdown(f'<div style="font-size:0.8rem;color:#7A7268;padding-top:0.4rem;">🚫 <b>{nm}</b> · {len(excluded[nm])}개 자료</div>', unsafe_allow_html=True)
                        with ec2:
                            if st.button("복원", key=f"auto_restore_{nm}", use_container_width=True):
                                st.session_state["auto_excluded"].discard(nm)
                                st.rerun()

            if unmatched:
                with st.expander(f"⚠️ 이름 인식 실패 {len(unmatched)}개 (분석 제외됨)"):
                    for f in unmatched:
                        st.markdown(f"- {f.name}")
                    st.caption("파일명에 명부상의 이름이 정확히 포함되도록 바꿔서 다시 올리면 인식됩니다.")

            st.markdown('<div style="height:0.7rem;"></div>', unsafe_allow_html=True)
            if st.button(f"◈  분석 시작 ({len(names_sorted)}명 일괄 분석)", use_container_width=True, key="auto_run"):
                if not names_sorted:
                    st.error("⚠️ 이름이 인식된 대상자가 없습니다. 파일명을 확인해주세요.")
                else:
                    st.session_state["auto_results"] = []
                    n_t = len(names_sorted)
                    prog = st.progress(0, text="분석 준비 중...")
                    for k, nm in enumerate(names_sorted):
                        prog.progress(k / n_t, text=f"({k+1}/{n_t}) {nm} 분석 중...")
                        try:
                            fd = {f.name: read_file_content(f) for f in grouped[nm]}
                            dept = _deptA.get(nm, "")
                            if dept: fd["소속 부서"] = dept
                            fd["회사 인재상"] = company_standard
                            fd["회사 5대 핵심문화 축"] = core_culture
                            fd["회사 향후 방향성"] = company_direction
                            R = analyze_candidate(api_key, fd, nm, company_standard)
                            R["material_coverage"] = sorted(coverage_from_filenames([f.name for f in grouped[nm]]))
                            st.session_state["auto_results"].append({"name": nm, "dept": dept, "result": R, "success": True})
                            save_to_archive({"saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                             "candidate_name": nm, "dept": dept, "result": R})
                        except Exception as e:
                            st.session_state["auto_results"].append({"name": nm, "dept": _deptA.get(nm, ""), "result": {}, "success": False, "error": str(e)})
                        prog.progress((k + 1) / n_t, text=f"({k+1}/{n_t}) {nm} 완료")
                    prog.progress(1.0, text="✅ 분석 완료!")
                    st.session_state["auto_done"] = True
                    st.rerun()
        elif not st.session_state["auto_done"]:
            st.markdown('<p style="font-size:0.82rem;color:#B0A898;text-align:center;padding:1.5rem 0;">위에서 자료를 올리면 파일명으로 대상자를 자동 인식합니다.</p>', unsafe_allow_html=True)

        # ── 자동 분석 결과 ──
        if st.session_state["auto_done"] and st.session_state["auto_results"]:
            st.markdown("""
            <div class="section-header" style="margin-top:1.5rem;">
                <span class="section-num">03</span>
                <span class="section-title">자동 분석 결과</span>
                <div class="section-rule"></div>
            </div>
            """, unsafe_allow_html=True)

            _dimk = ["cognitive_ability", "job_expertise", "proactiveness", "leadership"]
            _dimn = ["인지", "전문성", "적극성", "리더십"]
            _re_e = {"LOW": "🟢", "MEDIUM": "🟡", "HIGH": "🔴", "CRITICAL": "🔴"}.get
            _ve = {"KEEP": "✅", "DEVELOP": "📈", "WATCH": "👁", "MISFIT": "⚠️"}.get
            _scored = []
            for br in st.session_state["auto_results"]:
                ov = compute_overall_score(br["result"])[0] if br["success"] else None
                _scored.append((ov if ov is not None else -1, br, ov))
            _scored.sort(key=lambda x: x[0], reverse=True)
            _hdr = "| 순위 | 종합 | 이름 | 부서 | 판정 | " + " | ".join(_dimn) + " | 번아웃 | 이직 |"
            _nc = _hdr.count("|") - 1
            _rows = [_hdr, "|" + "|".join([":---:"] * _nc) + "|"]
            _rk = 0
            for _, br, ov in _scored:
                if not br["success"]:
                    _rows.append(f"| - | - | {br['name']} | {br['dept']} | ❌ | | | | | | |")
                    continue
                _rk += 1
                _md = {1: "🥇", 2: "🥈", 3: "🥉"}.get(_rk, str(_rk))
                R = br["result"]
                _dm = R.get("dimensions", {})
                _sc = [str(_dm.get(k, {}).get("score", "?")) for k in _dimk]
                _bl = R.get("burnout_risk", {}).get("level", "?")
                _tl = R.get("turnover_risk", {}).get("level", "?")
                _vd = R.get("rebalancing_verdict", {}).get("decision", "?")
                _rows.append(f"| {_md} | **{ov}** | **{br['name']}** | {br['dept']} | {_ve(_vd,'⚪')} {_vd} | {' | '.join(_sc)} | {_re_e(_bl,'⚪')} | {_re_e(_tl,'⚪')} |")
            st.markdown("\n".join(_rows))
            st.markdown('<p style="font-size:0.7rem;color:#B0A898;margin-top:0.5rem;">종합 점수 높은 순 · 판정: ✅ KEEP · 📈 DEVELOP · 👁 WATCH · ⚠️ MISFIT</p>', unsafe_allow_html=True)

            st.markdown("""
            <div class="section-header" style="margin-top:1.5rem;">
                <span class="section-num">04</span>
                <span class="section-title">개별 상세 결과</span>
                <div class="section-rule"></div>
            </div>
            """, unsafe_allow_html=True)
            for br in st.session_state["auto_results"]:
                icon = "✅" if br["success"] else "❌"
                with st.expander(f"{icon}  {br['name']} · {br.get('dept','')}"):
                    if br["success"]:
                        render_result(br["result"], br["name"])
                    else:
                        st.error(f"분석 실패: {br.get('error','')}")

            if st.button("🔄  초기화 (새 자동 분석 시작)", use_container_width=True, key="auto_reset"):
                st.session_state["auto_results"] = []
                st.session_state["auto_done"] = False
                st.session_state["auto_excluded"] = set()
                st.rerun()



# ════════════════════════════════════════════════════════
#  TAB 3 — 조직도
# ════════════════════════════════════════════════════════
with tab_org:
    st.markdown("""
    <div class="section-header">
        <span class="section-num">🏢</span>
        <span class="section-title">조직도 · Talent Status Map</span>
        <div class="section-rule"></div>
    </div>
    """, unsafe_allow_html=True)

    # 범례
    st.markdown("""
    <div style="background:white;border:1px solid #D4CEC4;border-radius:8px;
                padding:0.9rem 1.3rem;margin-bottom:1.2rem;display:flex;
                flex-wrap:wrap;gap:1.2rem;align-items:center;">
        <span style="font-size:0.72rem;font-weight:700;color:#3D3830;letter-spacing:1px;">상태 신호등</span>
        <span style="font-size:0.75rem;color:#7A7268;"><span style="display:inline-block;width:9px;height:9px;border-radius:50%;background:#2D6A4F;margin-right:4px;"></span>정상</span>
        <span style="font-size:0.75rem;color:#7A7268;"><span style="display:inline-block;width:9px;height:9px;border-radius:50%;background:#E0A800;margin-right:4px;"></span>번아웃 초기</span>
        <span style="font-size:0.75rem;color:#7A7268;"><span style="display:inline-block;width:9px;height:9px;border-radius:50%;background:#C0392B;margin-right:4px;"></span>집중관리 필요</span>
        <span style="font-size:0.75rem;color:#7A7268;"><span style="display:inline-block;width:9px;height:9px;border-radius:50%;background:#C0BCB4;margin-right:4px;"></span>자료 미등록</span>
        <span style="font-size:0.72rem;color:#B0A898;margin-left:0.4rem;">| 카드 왼쪽 굵은 테두리 = 부서장</span>
    </div>
    """, unsafe_allow_html=True)

    org_data = load_org_data()

    if not org_data:
        st.warning("⚠️ org_data.json 파일이 없습니다. GitHub에 org_data.json을 함께 업로드해주세요.")
    else:
        # 아카이브를 이름 기준 딕셔너리로 (최신 우선)
        archive_all = load_archive()
        archive_by_name = {}
        for rec in archive_all:
            nm = rec.get("candidate_name", "")
            if nm and nm not in archive_by_name:
                archive_by_name[nm] = rec

        # 통계
        total_people = sum(len(people) for div in org_data.values()
                           for team in div.values() for people in team.values())
        analyzed = sum(1 for div in org_data.values() for team in div.values()
                       for people in team.values() for p in people
                       if p["name"] in archive_by_name)

        st.markdown(f"""
        <div style="display:flex;gap:1rem;margin-bottom:1.5rem;">
            <div style="flex:1;background:white;border:1px solid #D4CEC4;border-radius:8px;padding:1rem;text-align:center;">
                <div style="font-family:'DM Serif Display',serif;font-size:1.8rem;font-style:italic;color:#1A1714;">{total_people}</div>
                <div style="font-size:0.62rem;letter-spacing:1px;text-transform:uppercase;color:#B0A898;">전체 인원</div>
            </div>
            <div style="flex:1;background:white;border:1px solid #D4CEC4;border-radius:8px;padding:1rem;text-align:center;">
                <div style="font-family:'DM Serif Display',serif;font-size:1.8rem;font-style:italic;color:#2D6A4F;">{analyzed}</div>
                <div style="font-size:0.62rem;letter-spacing:1px;text-transform:uppercase;color:#B0A898;">분석 완료</div>
            </div>
            <div style="flex:1;background:white;border:1px solid #D4CEC4;border-radius:8px;padding:1rem;text-align:center;">
                <div style="font-family:'DM Serif Display',serif;font-size:1.8rem;font-style:italic;color:#B0A898;">{total_people - analyzed}</div>
                <div style="font-size:0.62rem;letter-spacing:1px;text-transform:uppercase;color:#B0A898;">미분석</div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

        # ── 인터랙티브 조직도 (줌/팬 + 인앱 모달) ──
        st.markdown("""
        <p style="font-size:0.78rem;color:#7A7268;margin-bottom:0.5rem;">
        💡 <b>Ctrl + 휠</b> 마우스 위치 기준 확대/축소 &nbsp;·&nbsp; <b>Shift + 휠</b> 좌우 이동 &nbsp;·&nbsp; <b>드래그</b> 이동 &nbsp;·&nbsp; <b>인원 카드 클릭</b> 시 분석 결과 즉시 표시 (새로고침 없음)
        </p>
        """, unsafe_allow_html=True)

        # 각 인원의 분석 요약 데이터를 JS로 전달 (클릭 시 모달용)
        person_data = {}
        for div in org_data.values():
            for team in div.values():
                for people in team.values():
                    for p in people:
                        nm = p["name"]
                        s, c, R = get_person_status(nm, archive_by_name)
                        if R:
                            dims = R.get("dimensions", {})
                            person_data[nm] = {
                                "has": True,
                                "summary": R.get("candidate_summary", ""),
                                "verdict": R.get("rebalancing_verdict", {}).get("decision", "-"),
                                "confidence": R.get("rebalancing_verdict", {}).get("confidence", "-"),
                                "rationale": R.get("rebalancing_verdict", {}).get("rationale", ""),
                                "org_fit": R.get("org_fit", {}).get("score", "-"),
                                "leadership": R.get("leadership_readiness", {}).get("score", "-"),
                                "lead_rec": R.get("leadership_readiness", {}).get("recommendation", ""),
                                "cog": dims.get("cognitive_ability", {}).get("score", "-"),
                                "job": dims.get("job_expertise", {}).get("score", "-"),
                                "pro": dims.get("proactiveness", {}).get("score", "-"),
                                "lead": dims.get("leadership", {}).get("score", "-"),
                                "burnout": R.get("burnout_risk", {}).get("level", "-"),
                                "turnover": R.get("turnover_risk", {}).get("level", "-"),
                                "tags": R.get("personality_tags", [])[:5],
                                "insight": R.get("overall_insight", ""),
                                "overall": compute_overall_score(R)[0] if compute_overall_score(R)[0] is not None else "-",
                                "direction": R.get("direction_fit", {}).get("score", "-"),
                                "materials": R.get("data_coverage", {}).get("materials", []),
                                "cov_conf": R.get("data_coverage", {}).get("confidence", ""),
                                "cov_note": R.get("data_coverage", {}).get("note", ""),
                                "major": R.get("profile", {}).get("major", ""),
                                "univ": R.get("profile", {}).get("university", ""),
                                "edu": R.get("profile", {}).get("education_level", ""),
                                "region": R.get("profile", {}).get("region", ""),
                                "mat_cov": R.get("material_coverage", []),
                            }
                        else:
                            person_data[nm] = {"has": False}

        # ── 계층 트리(top-down) HTML 빌드 ──
        ACCENT = {"전략기획본부":"#3D4E8C","영업마케팅본부":"#1F7A6B","R&D본부":"#7A4E8C",
                  "품질인허가본부":"#B07A1E","생산본부":"#2E6B8C"}
        STATUS_LIGHT = {"none":"#C0BCB4","green":"#2D6A4F","yellow":"#E0A800","red":"#C0392B"}
        STATUS_LABEL = {"none":"자료 미등록","green":"정상","yellow":"번아웃 초기","red":"집중관리 필요"}

        def _is_leader(role):
            return bool(role) and (role in ("본부장","팀장","파트장","센터장","조장") or role.endswith("장"))

        def _cnt(node):
            if isinstance(node, list):
                return sum((1 if (isinstance(x, dict) and "name" in x) else _cnt(x)) for x in node)
            if isinstance(node, dict):
                return sum(_cnt(v) for v in node.values())
            return 0

        def _light_span(nm):
            s = get_person_status(nm, archive_by_name)[0]
            col = STATUS_LIGHT.get(s, "#C0BCB4")
            return s, col, STATUS_LABEL.get(s, "자료 미등록"), \
                   f'<span class="light" style="background:{col};box-shadow:0 0 5px {col};"></span>'

        def _person_card(p, accent):
            nm = p.get("name", ""); pos = p.get("pos", ""); role = p.get("role", "")
            _s, _col, label, light = _light_span(nm)
            role_txt = role if (role and role != "팀원") else ""
            meta = pos + ((" · " + role_txt) if role_txt else "")
            cls = "pcard leader" if _is_leader(role) else "pcard"
            safe = nm.replace("\\", "").replace("'", "")
            return (f'<div class="{cls}" style="--accent:{accent};" onclick="showModal(\'{safe}\')" title="{nm} · {label}">'
                    f'<div class="pname">{nm}</div>'
                    f'<div class="pmeta"><span class="ppos">{meta}</span>{light}</div></div>')

        def _people_grid(people, accent):
            if not people:
                return ""
            return '<div class="people">' + "".join(_person_card(p, accent) for p in people) + '</div>'

        def _hcard(kind, name, count, accent):
            return (f'<div class="hcard {kind}" style="--accent:{accent};">'
                    f'<div class="hname">{name}</div>'
                    f'<div class="hmeta"><span class="hkind">{kind}</span><span class="hcount">{count}명</span></div>'
                    f'</div>')

        def _part_li(pname, plist, accent):
            node = _hcard("파트", pname, _cnt(plist), accent) + _people_grid(plist, accent)
            return f'<li class="leaf"><div class="node">{node}</div></li>'

        def _team_li(tname, tdict, accent):
            direct = tdict.get("_직속", [])
            parts = [(k, v) for k, v in tdict.items() if k != "_직속"]
            node = _hcard("팀", tname, _cnt(tdict), accent) + _people_grid(direct, accent)
            if parts:
                kids = "".join(_part_li(k, v, accent) for k, v in parts)
                return f'<li><div class="node">{node}</div><ul>{kids}</ul></li>'
            return f'<li class="leaf"><div class="node">{node}</div></li>'

        def _bonbu_li(bname, bdict):
            accent = ACCENT.get(bname, "#5A5A5A")
            bd = bdict.get("_직속", {})
            direct = bd.get("_직속", []) if isinstance(bd, dict) else []
            teams = [(k, v) for k, v in bdict.items() if k != "_직속"]
            node = _hcard("본부", bname, _cnt(bdict), accent) + _people_grid(direct, accent)
            kids = "".join(_team_li(k, v, accent) for k, v in teams)
            return f'<li class="bonbu"><div class="node">{node}</div><ul>{kids}</ul></li>'

        # 대표 노드
        _ceo = next(iter((org_data.get("이사회", {}).get("_직속", {}) or {}).get("_직속", [])), None)
        _ceo_name = _ceo.get("name", "대표") if _ceo else "대표"
        _ceo_pos = ((_ceo.get("pos", "") + " · " + _ceo.get("role", "")).strip(" ·")) if _ceo else ""
        _total = _cnt(org_data)
        _cs, _cc, _clabel, _clight = _light_span(_ceo_name)
        ceo_card = (f'<div class="hcard ceo" style="--accent:#2B2F36;" onclick="showModal(\'{_ceo_name}\')" '
                    f'title="{_ceo_name} · {_clabel}">'
                    f'<div class="hname">{_ceo_name}{_clight}</div>'
                    f'<div class="hmeta"><span class="hkind">대표이사</span><span class="hcount">{_total}명</span></div>'
                    f'</div>')
        _bonbus = [b for b in org_data.keys() if b != "이사회"]
        org_html = ('<ul class="tree-root"><li><div class="node">' + ceo_card + '</div><ul>'
                    + "".join(_bonbu_li(b, org_data[b]) for b in _bonbus)
                    + '</ul></li></ul>')
        pdata_json = json.dumps(person_data, ensure_ascii=False)

        org_chart_html = """
        <!DOCTYPE html><html><head><meta charset="utf-8">
        <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;600;700;800&display=swap" rel="stylesheet">
        <style>
        * { box-sizing:border-box; margin:0; padding:0; }
        body { font-family:'Noto Sans KR',sans-serif; background:#F7F3ED; overflow:hidden; }
        #viewport { width:100%; height:840px; overflow:hidden; position:relative;
                    background:#F2EDE5; border:1px solid #D4CEC4; border-radius:12px; cursor:grab; }
        #viewport.grabbing { cursor:grabbing; }
        #canvas { transform-origin:0 0; padding:36px; display:inline-block; }
        /* 트리 컨테이너 */
        .tree-root, li > ul { display:flex; justify-content:center; padding-top:26px; position:relative; margin:0; list-style:none; }
        .tree-root { padding-top:0; }
        li { list-style:none; display:flex; flex-direction:column; align-items:center; position:relative; padding:26px 10px 0; }
        .tree-root > li { padding-top:0; }
        li::before, li::after { content:''; position:absolute; top:0; right:50%; border-top:1.5px solid #C9C1B4; width:50%; height:26px; }
        li::after { right:auto; left:50%; border-left:1.5px solid #C9C1B4; }
        li:only-child::before, li:only-child::after { display:none; }
        li:only-child { padding-top:26px; }
        .tree-root > li:only-child { padding-top:0; }
        li:first-child::before, li:last-child::after { border:0; }
        li:last-child::before { border-right:1.5px solid #C9C1B4; }
        li > ul::before { content:''; position:absolute; top:0; left:50%; border-left:1.5px solid #C9C1B4; height:26px; }
        .node { display:flex; flex-direction:column; align-items:center; }
        /* ── 모든 카드 동일 크기 (디자인 통일) ── */
        .hcard, .pcard {
            width:148px; box-sizing:border-box; background:#FFFFFF; border:1px solid #DDD7CC;
            border-radius:8px; padding:8px 11px; cursor:pointer; text-align:left;
            display:flex; flex-direction:column; justify-content:center; gap:3px; min-height:52px;
            box-shadow:0 1px 3px rgba(0,0,0,0.05);
            transition:transform .12s, box-shadow .12s, border-color .12s;
        }
        .hcard:hover, .pcard:hover { border-color:var(--accent); box-shadow:0 4px 12px rgba(0,0,0,0.13); transform:translateY(-2px); }
        /* 구조 헤더 카드: 상단 accent 라인으로 구분 */
        .hcard { border-top:3px solid var(--accent); }
        .hcard .hname { font-size:12.5px; font-weight:800; color:#1A1714; line-height:1.25;
                        display:flex; align-items:center; gap:6px; }
        .hcard.본부 .hname { color:var(--accent); }
        .hcard.ceo .hname { color:#2B2F36; }
        .hcard .hmeta { display:flex; align-items:center; justify-content:space-between; gap:6px; }
        .hcard .hkind { font-size:9px; letter-spacing:1px; text-transform:uppercase; color:var(--accent); font-weight:800; }
        .hcard .hcount { font-size:9.5px; color:#8A8278; }
        .hcard .light { width:9px; height:9px; border-radius:50%; flex:0 0 auto; }
        /* 인원 카드: 부서장은 왼쪽 굵은 테두리로 구분 */
        .pcard.leader { border-left:4px solid var(--accent); background:#FCFBF8; }
        .pname { font-size:12.5px; font-weight:600; color:#1A1714; line-height:1.25; }
        .pcard.leader .pname { font-weight:800; }
        .pmeta { display:flex; align-items:center; justify-content:space-between; gap:7px; }
        .ppos { font-size:10px; color:#9A938A; }
        .pcard.leader .ppos { color:var(--accent); font-weight:600; }
        .light { width:9px; height:9px; border-radius:50%; flex:0 0 auto; }
        /* 인원 그리드 */
        .people { display:flex; flex-wrap:wrap; justify-content:center; gap:8px; max-width:336px; margin-top:13px; }
        #controls { position:absolute; bottom:16px; right:16px; display:flex; gap:7px; z-index:10; }
        #controls button { width:38px; height:38px; border:1px solid #D4CEC4; background:#FFFFFF;
                           border-radius:8px; font-size:18px; cursor:pointer; color:#3D3830;
                           box-shadow:0 2px 7px rgba(0,0,0,0.1); }
        #controls button:hover { background:#EDE8E0; }
        #zoomlabel { position:absolute; bottom:16px; left:16px; background:#FFFFFF;
                     border:1px solid #D4CEC4; border-radius:8px; padding:6px 14px;
                     font-size:12px; color:#7A7268; z-index:10; }
        /* 모달 */
        #overlay { display:none; position:fixed; inset:0; background:rgba(26,23,20,0.55);
                   z-index:100; align-items:center; justify-content:center; }
        #modal { background:#F7F3ED; width:min(720px,92%); max-height:88%; overflow-y:auto;
                 border-radius:14px; padding:0; box-shadow:0 20px 60px rgba(0,0,0,0.3); }
        .m-head { padding:1.6rem 2rem 1.2rem; border-bottom:1px solid #E2DDD4; position:relative; }
        .m-name { font-size:1.6rem; font-weight:800; color:#1A1714; }
        .m-summary { font-size:0.9rem; color:#7A7268; margin-top:0.4rem; line-height:1.6; }
        .m-close { position:absolute; top:1.3rem; right:1.5rem; cursor:pointer; font-size:1.3rem;
                   color:#B0A898; border:none; background:none; }
        .m-close:hover { color:#1A1714; }
        .m-body { padding:1.5rem 2rem 2rem; }
        .m-verdict { display:flex; align-items:center; justify-content:space-between;
                     border-radius:10px; padding:1rem 1.4rem; margin-bottom:1.2rem; }
        .m-tags { margin:0.3rem 0 0; }
        .m-tag { display:inline-block; border:1px solid #D4AF72; color:#B8924A; border-radius:4px;
                 padding:2px 9px; font-size:0.72rem; margin:2px; }
        .m-grid { display:grid; grid-template-columns:1fr 1fr; gap:0.7rem; margin-bottom:1.2rem; }
        .m-metric { background:#FFFFFF; border:1px solid #D4CEC4; border-radius:8px; padding:0.9rem 1.1rem; }
        .m-metric .lbl { font-size:0.62rem; letter-spacing:1px; text-transform:uppercase; color:#B0A898; }
        .m-metric .vl { font-size:1.5rem; font-weight:800; font-family:'Noto Sans KR'; }
        .m-row { display:flex; gap:0.5rem; margin-bottom:1.2rem; }
        .m-chip { flex:1; background:#FFFFFF; border:1px solid #D4CEC4; border-radius:8px;
                  padding:0.7rem; text-align:center; }
        .m-chip .lbl { font-size:0.6rem; color:#B0A898; text-transform:uppercase; }
        .m-chip .vl { font-size:0.95rem; font-weight:700; margin-top:0.2rem; }
        .m-sec { background:#FFFFFF; border:1px solid #D4CEC4; border-radius:8px; padding:1rem 1.3rem;
                 margin-bottom:0.9rem; }
        .m-sec .h { font-size:0.65rem; font-weight:700; letter-spacing:1.5px; text-transform:uppercase;
                    color:#B8924A; margin-bottom:0.5rem; }
        .m-sec .t { font-size:0.85rem; color:#3D3830; line-height:1.75; }
        .m-empty { text-align:center; padding:2.5rem 1rem; }
        .m-empty .ic { font-size:2.5rem; }
        .m-empty .ttl { font-size:1.1rem; font-weight:700; color:#8B6914; margin:0.8rem 0 0.4rem; }
        .m-empty .ds { font-size:0.85rem; color:#7A7268; line-height:1.6; }
        .light { cursor:pointer; }
        .pcard:hover .light { transform:scale(1.45); transition:transform .12s; }
        .m-note { font-size:0.78rem; color:#7A7268; line-height:1.6; margin-top:0.5rem;
                  padding:0.7rem 1rem; background:#F2EEE6; border-radius:8px; }
        </style></head><body>
        <div id="viewport">
            <div id="canvas">__ORG_HTML__</div>
            <div id="zoomlabel">70%</div>
            <div id="controls">
                <button onclick="zoomBtn(0.1)">+</button>
                <button onclick="zoomBtn(-0.1)">−</button>
                <button onclick="resetView()">⊙</button>
            </div>
        </div>
        <div id="overlay" onclick="if(event.target===this)closeModal()">
            <div id="modal"></div>
        </div>
        <script>
        const PDATA = __PDATA__;
        const STD_MAT = __STDMAT__;
        let scale=0.7, tx=0, ty=0, panning=false, sx=0, sy=0;
        const vp=document.getElementById('viewport');
        const cv=document.getElementById('canvas');
        const zl=document.getElementById('zoomlabel');
        function apply(){ cv.style.transform='translate('+tx+'px,'+ty+'px) scale('+scale+')';
                          zl.textContent=Math.round(scale*100)+'%'; }
        function zoomBtn(d){ scale=Math.min(2.5,Math.max(0.25,scale+d)); apply(); }
        function resetView(){ scale=0.7; tx=0; ty=0; apply(); }
        vp.addEventListener('wheel',function(e){
            if(e.ctrlKey){ e.preventDefault();
                const r=vp.getBoundingClientRect();
                const mx=e.clientX-r.left, my=e.clientY-r.top, before=scale;
                scale=Math.min(2.5,Math.max(0.25,scale - e.deltaY*0.0015));
                tx=mx-(mx-tx)*(scale/before); ty=my-(my-ty)*(scale/before); apply();
            } else if(e.shiftKey){ e.preventDefault();
                tx -= (e.deltaY || e.deltaX); apply();
            }
        },{passive:false});
        vp.addEventListener('mousedown',function(e){
            if(e.target.closest('.pcard'))return;
            panning=true; sx=e.clientX-tx; sy=e.clientY-ty; vp.classList.add('grabbing');
        });
        window.addEventListener('mousemove',function(e){ if(panning){ tx=e.clientX-sx; ty=e.clientY-sy; apply(); }});
        window.addEventListener('mouseup',function(){ panning=false; vp.classList.remove('grabbing'); });

        const V_STYLE = {
            "KEEP":   ["#2D6A4F","#EAF4EE","✓ KEEP — 핵심 인재, 유지 권장"],
            "DEVELOP":["#2B3D5C","#E8EEF5","↗ DEVELOP — 육성 대상"],
            "WATCH":  ["#8B6914","#FBF3E0","◷ WATCH — 관찰 필요"],
            "MISFIT": ["#8B2635","#FAEAEC","✕ MISFIT — 조직 방향성 부적합"]
        };
        const RISK = {"LOW":"🟢 낮음","MEDIUM":"🟡 주의","HIGH":"🔴 높음","CRITICAL":"🔴 심각"};
        function showModal(name){
            const d=PDATA[name]; const m=document.getElementById('modal');
            if(!d || !d.has){
                m.innerHTML='<div class="m-head"><button class="m-close" onclick="closeModal()">✕</button>'
                  +'<div class="m-name">'+name+'</div></div>'
                  +'<div class="m-body"><div class="m-empty"><div class="ic">📭</div>'
                  +'<div class="ttl">분석 결과가 없습니다 · 자료 미등록</div>'
                  +'<div class="ds">'+name+' 님은 아직 분석 데이터가 없습니다.<br><b>‘개인 분석’ 탭</b>에서 이 직원을 선택해 검사를 진행하세요.</div>'
                  +'</div></div>';
            } else {
                const vs=V_STYLE[d.verdict]||["#7A7268","#EDE8E0",d.verdict];
                let tags=d.tags.map(t=>'<span class="m-tag">'+t+'</span>').join('');
                m.innerHTML='<div class="m-head"><button class="m-close" onclick="closeModal()">✕</button>'
                  +'<div class="m-name">'+name+'</div>'
                  +'<div class="m-summary">'+(d.summary||'')+'</div>'
                  +'<div class="m-tags">'+tags+'</div></div>'
                  +'<div class="m-body">'
                  +'<div class="m-verdict" style="background:'+vs[1]+';border:2px solid '+vs[0]+';">'
                  +'<div><div style="font-size:0.6rem;letter-spacing:2px;text-transform:uppercase;color:'+vs[0]+';">Rebalancing Verdict</div>'
                  +'<div style="font-size:1.05rem;font-weight:800;color:'+vs[0]+';">'+vs[2]+'</div></div>'
                  +'<div style="text-align:right;"><div style="font-size:0.6rem;color:#7A7268;">신뢰도</div>'
                  +'<div style="font-weight:700;color:'+vs[0]+';">'+d.confidence+'</div></div></div>'
                  +'<div style="background:#FFFFFF;border:2px solid #2B3D5C;border-radius:10px;padding:0.7rem 1.1rem;margin-bottom:1.2rem;display:flex;align-items:center;justify-content:space-between;">'
                  +'<div style="font-size:0.6rem;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:#2B3D5C;">종합 점수</div>'
                  +'<div style="font-size:1.7rem;font-weight:800;color:#2B3D5C;">'+d.overall+'<span style="font-size:0.7rem;color:#B0A898;font-weight:400;">/100</span></div>'
                  +'</div>'
                  +'<div class="m-grid">'
                  +'<div class="m-metric"><div class="lbl">조직 적합도</div><div class="vl" style="color:'+(d.org_fit>=70?'#2D6A4F':d.org_fit>=50?'#8B6914':'#8B2635')+';">'+d.org_fit+'<span style="font-size:0.7rem;color:#B0A898;">/100</span></div></div>'
                  +'<div class="m-metric"><div class="lbl">리더십 준비도</div><div class="vl" style="color:'+(d.leadership>=80?'#2D6A4F':d.leadership>=60?'#2B3D5C':'#8B2635')+';">'+d.leadership+'<span style="font-size:0.7rem;color:#B0A898;">/100</span></div></div>'
                  +'<div class="m-metric"><div class="lbl">방향성 적합도</div><div class="vl" style="color:'+(d.direction>=70?'#2D6A4F':d.direction>=50?'#8B6914':'#8B2635')+';">'+d.direction+'<span style="font-size:0.7rem;color:#B0A898;">/100</span></div></div>'
                  +'</div>'
                  +'<div class="m-row">'
                  +'<div class="m-chip"><div class="lbl">인지</div><div class="vl">'+d.cog+'</div></div>'
                  +'<div class="m-chip"><div class="lbl">전문성</div><div class="vl">'+d.job+'</div></div>'
                  +'<div class="m-chip"><div class="lbl">적극성</div><div class="vl">'+d.pro+'</div></div>'
                  +'<div class="m-chip"><div class="lbl">리더십</div><div class="vl">'+d.lead+'</div></div>'
                  +'</div>'
                  +'<div class="m-row">'
                  +'<div class="m-chip"><div class="lbl">번아웃</div><div class="vl">'+(RISK[d.burnout]||d.burnout)+'</div></div>'
                  +'<div class="m-chip"><div class="lbl">이직 위험</div><div class="vl">'+(RISK[d.turnover]||d.turnover)+'</div></div>'
                  +'</div>'
                  +'<div class="m-sec"><div class="h">리밸런싱 판단 근거</div><div class="t">'+(d.rationale||'')+'</div></div>'
                  +'<div class="m-sec"><div class="h">리더 부여 결론</div><div class="t">'+(d.lead_rec||'')+'</div></div>'
                  +'<div class="m-sec"><div class="h">종합 인사이트</div><div class="t">'+(d.insight||'')+'</div></div>'
                  +'<div class="m-sec"><div class="h">제출 자료 · 서류 리스트</div><div class="t">'
                    +((d.materials&&d.materials.length)?d.materials.map(x=>'<span class="m-tag">'+x+'</span>').join(''):'<span style="color:#B0A898;">기록된 제출 자료 없음</span>')
                    +(d.cov_conf?'<div style="margin-top:7px;font-size:0.78rem;color:#7A7268;">근거 신뢰도 <b>'+d.cov_conf+'</b>'+(d.cov_note?' · '+d.cov_note:'')+'</div>':'')
                    +'</div></div>'
                  +'<div class="m-sec"><div class="h">8대 표준 분석자료 충족 현황 ('+((d.mat_cov||[]).length)+'/8)</div><div class="t">'
                    +STD_MAT.map(function(x){var on=(d.mat_cov||[]).indexOf(x)>=0;return '<span style="display:inline-flex;align-items:center;gap:4px;border-radius:6px;padding:3px 9px;margin:2px;font-size:0.76rem;background:'+(on?'#E6F2EA':'#F2EEE6')+';border:1px solid '+(on?'#9CCBB0':'#E2DDD4')+';color:'+(on?'#1E5C3A':'#B0A898')+';font-weight:'+(on?'700':'500')+';">'+(on?'✅':'⚪')+' '+x+'</span>';}).join('')
                    +'</div></div>'
                  +'<div class="m-sec"><div class="h">프로필</div><div class="t">'
                    +(function(){var a=[['전공',d.major],['대학',d.univ],['학력',d.edu],['출신지역',d.region]].filter(x=>x[1]&&x[1]!='자료 미제공'&&x[1]!='-');return a.length?a.map(x=>'<b>'+x[0]+'</b> '+x[1]).join(' &nbsp;·&nbsp; '):'<span style="color:#B0A898;">프로필 자료 미확인</span>';})()
                    +'</div></div>'
                  +'<div class="m-note">자료를 추가해 다시 검사하려면 <b>‘개인 분석’ 탭</b>에서 이 직원을 선택해 진행하세요.</div>'
                  +'</div>';
            }
            document.getElementById('overlay').style.display='flex';
        }
        function closeModal(){ document.getElementById('overlay').style.display='none'; }
        apply();
        </script>
        </body></html>
        """
        org_chart_html = org_chart_html.replace("__ORG_HTML__", org_html).replace("__PDATA__", pdata_json).replace("__STDMAT__", json.dumps(STANDARD_MATERIALS, ensure_ascii=False))
        components.html(org_chart_html, height=870, scrolling=False)



# ── Footer ──
st.markdown("""
<div style="text-align:center;padding:3rem 0 1.5rem;font-size:0.65rem;
     letter-spacing:2px;text-transform:uppercase;color:#C8C0B4;
     border-top:1px solid #E2DDD4;margin-top:3rem;">
    Talent Intelligence Platform &nbsp;·&nbsp; M.I.Tech P&C Team &nbsp;·&nbsp; Powered by Claude AI
</div>
""", unsafe_allow_html=True)
